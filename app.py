from __future__ import annotations

import base64
import html
import os
import json
import hashlib
import mimetypes
import shutil
import subprocess
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO
from math import ceil
from pathlib import Path

import streamlit as st
import streamlit.components.v1 as components
from dotenv import load_dotenv
from googleapiclient.errors import HttpError
from openpyxl import load_workbook
from PIL import Image, ImageOps
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas
import xlrd

from src.drive_service import (
    DEFAULT_DRIVE_ACCOUNT,
    DriveService,
    get_setting,
    list_drive_configurations,
)
from src.local_archive import (
    load_local_archive_run_index,
    load_local_archive_run_index_for_date,
    regroup_direct_run_folders_by_date,
)
from src.models import ImageFile, ParseError, RunFolder
from src.parser import parse_run_folder_name
from src.ui_helpers import (
    group_runs_by_customer,
    render_parse_errors,
)

load_dotenv()

st.set_page_config(
    page_title="Drive Run Dashboard",
    page_icon=":open_file_folder:",
    layout="wide",
)

st.markdown(
    """
    <style>
    .customer-card {
        border: 1px solid #d6dde8;
        border-radius: 14px;
        padding: 1rem 1.1rem;
        margin: 0.75rem 0 0.35rem 0;
        background: linear-gradient(180deg, #ffffff 0%, #f7faff 100%);
        box-shadow: 0 1px 2px rgba(16, 24, 40, 0.04);
    }
    .customer-card-title {
        font-size: 1.05rem;
        font-weight: 700;
        color: #10233f;
        margin-bottom: 0.2rem;
    }
    .customer-card-meta {
        font-size: 0.92rem;
        color: #516074;
    }
    .customer-card-open {
        border-left: 4px solid #2b6df3;
        background: linear-gradient(180deg, #f8fbff 0%, #eef5ff 100%);
    }
    .run-shell {
        border: 1px solid #e3e8f0;
        border-radius: 12px;
        padding: 1rem;
        margin: 0.5rem 0 1rem 0;
        background: #ffffff;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

CACHE_DIR = Path(".cache")
RUN_DATA_CACHE_PATH = CACHE_DIR / "run_data.json"
RUN_DATA_CACHE_VERSION = "v2"
IMAGE_CACHE_DIR = CACHE_DIR / "images"
IMAGE_CACHE_VERSION = "v5"
INDEX_SYNC_STATUS_PATH = CACHE_DIR / "index_sync_status.json"
INDEX_SYNC_SCRIPT_PATH = Path(__file__).with_name("sync_index.py")
INDEX_SYNC_STALE_MINUTES = 120
INDEX_SYNC_UI_REFRESH_MS = 15000
USB_CAMERA_AUTO_ROTATE_CARRIERS = {"CargoSnapArchive"}
RECENT_REFRESH_DAYS_BEFORE = 1
RECENT_REFRESH_DAYS_AFTER = 0


@st.cache_resource(show_spinner=False)
def get_drive_service(account_name: str = DEFAULT_DRIVE_ACCOUNT) -> DriveService:
    return DriveService.from_service_account_env(account_name)


def _load_configured_google_runs(
    parse_errors: list[ParseError],
    archive_cutoff_date: date,
    maximum_date: date | None = None,
) -> list[RunFolder]:
    runs: list[RunFolder] = []
    for drive_config in list_drive_configurations():
        drive_service = get_drive_service(drive_config.account_name)
        runs.extend(
            _load_google_drive_run_index(
                drive_service=drive_service,
                root_folder_id=drive_config.root_folder_id,
                parse_errors=parse_errors,
                archive_cutoff_date=archive_cutoff_date,
                maximum_date=maximum_date,
                drive_account=drive_config.account_name,
            )
        )
    return runs


@st.cache_data(ttl=300, show_spinner=False)
def _load_run_data_payload(_refresh_token: int) -> dict[str, object]:
    cached_payload = _read_persisted_run_data()
    if cached_payload is not None:
        return _serialize_runs_payload(*cached_payload)

    runs: list[RunFolder] = []
    parse_errors: list[ParseError] = []
    archive_cutoff_days = int(get_setting("LOCAL_ARCHIVE_AFTER_DAYS") or "7")
    archive_cutoff_date = date.today() - timedelta(days=archive_cutoff_days)
    google_runs: list[RunFolder] = []

    try:
        google_runs = _load_configured_google_runs(
            parse_errors=parse_errors,
            archive_cutoff_date=archive_cutoff_date,
        )
        runs.extend(google_runs)
    except RuntimeError:
        # Allow local archive-only usage if Drive is not configured.
        pass

    local_archive_root = get_setting("LOCAL_ARCHIVE_ROOT")
    if local_archive_root:
        archive_root = Path(local_archive_root)
        regroup_direct_run_folders_by_date(archive_root)
        local_runs = load_local_archive_run_index(
            archive_root=archive_root,
            parse_errors=parse_errors,
            archive_cutoff_date=None,
        )
        runs = _merge_runs_prefer_google_drive(runs, local_runs)

    runs = _mark_recent_runs_volatile(runs)
    parse_errors = _deduplicate_parse_errors(parse_errors)
    runs.sort(
        key=lambda run: (
            run.run_date,
            run.customer_code.lower(),
            run.carrier.lower(),
            run.run_id or "",
        )
    )
    _write_persisted_run_data(runs, parse_errors)
    return _serialize_runs_payload(runs, parse_errors)


def load_run_data(_refresh_token: int) -> tuple[list[RunFolder], list[ParseError]]:
    return _deserialize_runs_payload(_load_run_data_payload(_refresh_token))


def refresh_runs_for_date(
    selected_date: date,
    existing_runs: list[RunFolder],
    existing_parse_errors: list[ParseError],
) -> tuple[list[RunFolder], list[ParseError]]:
    runs: list[RunFolder] = [
        run for run in existing_runs if run.run_date != selected_date
    ]
    parse_errors: list[ParseError] = [
        error for error in existing_parse_errors if _parse_error_date(error) != selected_date
    ]

    fresh_parse_errors: list[ParseError] = []
    google_runs: list[RunFolder] = []

    try:
        google_runs = _load_configured_google_runs(
            parse_errors=fresh_parse_errors,
            archive_cutoff_date=selected_date,
            maximum_date=selected_date,
        )
        runs.extend(google_runs)
    except RuntimeError:
        pass

    local_archive_root = get_setting("LOCAL_ARCHIVE_ROOT")
    if local_archive_root:
        archive_root = Path(local_archive_root)
        regroup_direct_run_folders_by_date(archive_root)
        local_runs = load_local_archive_run_index_for_date(
            archive_root=archive_root,
            parse_errors=fresh_parse_errors,
            selected_date=selected_date,
        )
        runs = _merge_runs_prefer_google_drive(runs, local_runs)

    runs = _mark_recent_runs_volatile(runs)
    parse_errors = _deduplicate_parse_errors(parse_errors + fresh_parse_errors)
    runs.sort(
        key=lambda run: (
            run.run_date,
            run.customer_code.lower(),
            run.carrier.lower(),
            run.run_id or "",
        )
    )
    _write_persisted_run_data(runs, parse_errors)
    return runs, parse_errors


def _recent_refresh_window(today: date | None = None) -> tuple[date, date]:
    reference_day = today or date.today()
    return (
        reference_day - timedelta(days=RECENT_REFRESH_DAYS_BEFORE),
        reference_day + timedelta(days=RECENT_REFRESH_DAYS_AFTER),
    )


def _is_in_recent_refresh_window(run_date: date, today: date | None = None) -> bool:
    window_start, window_end = _recent_refresh_window(today=today)
    return window_start <= run_date <= window_end


def _mark_recent_runs_volatile(runs: list[RunFolder]) -> list[RunFolder]:
    for run in runs:
        run.metadata["volatile_recent"] = _is_in_recent_refresh_window(run.run_date)
    return runs


def _refresh_recent_cached_run_data(
    cached_payload: tuple[list[RunFolder], list[ParseError]],
) -> tuple[list[RunFolder], list[ParseError]]:
    cached_runs, cached_errors = cached_payload
    cached_runs = _mark_recent_runs_volatile(cached_runs)

    try:
        parse_errors: list[ParseError] = []
        recent_start, recent_end = _recent_refresh_window()
        recent_google_runs = _load_configured_google_runs(
            parse_errors=parse_errors,
            archive_cutoff_date=recent_start,
            maximum_date=recent_end,
        )
    except RuntimeError:
        return cached_runs, cached_errors

    recent_google_runs = _mark_recent_runs_volatile(recent_google_runs)

    retained_runs = [
        run
        for run in cached_runs
        if not (
            str(run.metadata.get("source")) == "google_drive"
            and _is_in_recent_refresh_window(run.run_date)
        )
    ]
    merged_runs = _merge_runs_prefer_google_drive(recent_google_runs, retained_runs)
    merged_runs.sort(
        key=lambda run: (
            run.run_date,
            run.customer_code.lower(),
            run.carrier.lower(),
            run.run_id or "",
        )
    )
    return merged_runs, _deduplicate_parse_errors(cached_errors + parse_errors)


def _deduplicate_parse_errors(parse_errors: list[ParseError]) -> list[ParseError]:
    unique_errors: dict[tuple[str, str, str, str], ParseError] = {}
    for error in parse_errors:
        unique_errors[
            (
                error.folder_id,
                error.folder_name,
                error.carrier,
                error.reason,
            )
        ] = error
    return list(unique_errors.values())


def _run_identity(run: RunFolder) -> tuple[str, str, date, str, str]:
    return (
        run.carrier.casefold(),
        run.customer_code.casefold(),
        run.run_date,
        (run.run_id or "").casefold(),
        run.folder_name.casefold(),
    )


def _merge_runs_prefer_google_drive(primary_runs: list[RunFolder], fallback_runs: list[RunFolder]) -> list[RunFolder]:
    merged: dict[tuple[str, str, date, str, str], RunFolder] = {
        _run_identity(run): run for run in fallback_runs
    }
    for run in primary_runs:
        merged[_run_identity(run)] = run
    return list(merged.values())


def _parse_error_date(error: ParseError) -> date | None:
    try:
        return parse_run_folder_name(error.folder_name).run_date
    except ValueError:
        return None


def _read_persisted_run_data() -> tuple[list[RunFolder], list[ParseError]] | None:
    if not RUN_DATA_CACHE_PATH.exists():
        return None

    try:
        payload = json.loads(RUN_DATA_CACHE_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError, ValueError):
        return None

    if payload.get("cache_version") != RUN_DATA_CACHE_VERSION:
        return None

    try:
        runs, parse_errors = _deserialize_runs_payload(payload)
    except (KeyError, TypeError, ValueError):
        return None

    return runs, parse_errors


def _read_persisted_run_cache_timestamp() -> datetime | None:
    if not RUN_DATA_CACHE_PATH.exists():
        return None

    try:
        payload = json.loads(RUN_DATA_CACHE_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError, ValueError):
        return None

    if payload.get("cache_version") != RUN_DATA_CACHE_VERSION:
        return None

    generated_at = payload.get("generated_at")
    if not isinstance(generated_at, str):
        return None

    try:
        return datetime.fromisoformat(generated_at)
    except ValueError:
        return None


def _read_index_sync_status() -> dict[str, object]:
    if not INDEX_SYNC_STATUS_PATH.exists():
        return {}

    try:
        payload = json.loads(INDEX_SYNC_STATUS_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError, ValueError):
        return {}

    if not isinstance(payload, dict):
        return {}
    return _expire_stale_index_sync_status(payload)


def _expire_stale_index_sync_status(payload: dict[str, object]) -> dict[str, object]:
    if payload.get("state") != "running":
        return payload

    reference_value = payload.get("updated_at") or payload.get("started_at")
    if not isinstance(reference_value, str):
        return payload

    try:
        reference_time = datetime.fromisoformat(reference_value)
    except ValueError:
        return payload

    now = datetime.now(reference_time.tzinfo) if reference_time.tzinfo else datetime.now()
    if now - reference_time <= timedelta(minutes=INDEX_SYNC_STALE_MINUTES):
        return payload

    stale_payload = dict(payload)
    stale_payload["state"] = "failed"
    stale_payload["error"] = (
        f"Background sync timed out after more than {INDEX_SYNC_STALE_MINUTES} minutes "
        "without finishing."
    )
    stale_payload["updated_at"] = datetime.now().astimezone().isoformat(timespec="seconds")
    try:
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        INDEX_SYNC_STATUS_PATH.write_text(
            json.dumps(stale_payload, ensure_ascii=True, indent=2),
            encoding="utf-8",
        )
    except OSError:
        return stale_payload
    return stale_payload


def _format_timestamp(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    try:
        return datetime.fromisoformat(value).astimezone().strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None


def _start_index_sync(mode: str, selected_date: date | None = None) -> bool:
    if not INDEX_SYNC_SCRIPT_PATH.exists():
        return False

    CACHE_DIR.mkdir(parents=True, exist_ok=True)

    command = [sys.executable, str(INDEX_SYNC_SCRIPT_PATH), "--mode", mode]
    if selected_date is not None:
        command.extend(["--date", selected_date.isoformat()])

    creationflags = 0
    if os.name == "nt":
        creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0) | getattr(
            subprocess,
            "DETACHED_PROCESS",
            0,
        )

    try:
        subprocess.Popen(
            command,
            cwd=str(Path(__file__).resolve().parent),
            close_fds=True,
            creationflags=creationflags,
        )
    except OSError:
        return False
    return True


def _render_background_sync_autorefresh(enabled: bool) -> None:
    if not enabled:
        return

    components.html(
        f"""
        <script>
          window.setTimeout(() => {{
            const parentWindow = window.parent;
            if (parentWindow && typeof parentWindow.location?.reload === "function") {{
              parentWindow.location.reload();
            }}
          }}, {INDEX_SYNC_UI_REFRESH_MS});
        </script>
        """,
        height=0,
    )


def _write_persisted_run_data(runs: list[RunFolder], parse_errors: list[ParseError]) -> None:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    payload = _serialize_runs_payload(runs, parse_errors)
    payload["cache_version"] = RUN_DATA_CACHE_VERSION
    payload["generated_at"] = datetime.now().astimezone().isoformat(timespec="seconds")
    RUN_DATA_CACHE_PATH.write_text(
        json.dumps(payload, ensure_ascii=True, indent=2),
        encoding="utf-8",
    )


def _serialize_runs_payload(
    runs: list[RunFolder],
    parse_errors: list[ParseError],
) -> dict[str, object]:
    return {
        "runs": [
            {
                "folder_id": run.folder_id,
                "folder_name": run.folder_name,
                "customer_code": run.customer_code,
                "run_date": run.run_date.isoformat(),
                "carrier": run.carrier,
                "run_id": run.run_id,
                "images": [
                    {
                        "id": image.id,
                        "name": image.name,
                        "mime_type": image.mime_type,
                        "web_view_link": image.web_view_link,
                        "size": image.size,
                    }
                    for image in run.images
                ],
                "qr_info": run.qr_info,
                "qr_source": run.qr_source,
                "metadata": dict(run.metadata),
            }
            for run in runs
        ],
        "parse_errors": [
            {
                "folder_id": error.folder_id,
                "folder_name": error.folder_name,
                "carrier": error.carrier,
                "reason": error.reason,
            }
            for error in parse_errors
        ],
    }


def _deserialize_runs_payload(
    payload: dict[str, object],
) -> tuple[list[RunFolder], list[ParseError]]:
    runs = [
        RunFolder(
            folder_id=item["folder_id"],
            folder_name=item["folder_name"],
            customer_code=item["customer_code"],
            run_date=date.fromisoformat(item["run_date"]),
            carrier=item["carrier"],
            run_id=item.get("run_id"),
            images=[
                ImageFile(
                    id=image["id"],
                    name=image["name"],
                    mime_type=image["mime_type"],
                    web_view_link=image.get("web_view_link"),
                    size=image.get("size"),
                )
                for image in item.get("images", [])
            ],
            qr_info=item.get("qr_info", "No QR info found"),
            qr_source=item.get("qr_source"),
            metadata=dict(item.get("metadata", {})),
        )
        for item in payload.get("runs", [])
    ]
    parse_errors = [
        ParseError(
            folder_id=item["folder_id"],
            folder_name=item["folder_name"],
            carrier=item["carrier"],
            reason=item["reason"],
        )
        for item in payload.get("parse_errors", [])
    ]
    return runs, parse_errors


def _image_cache_path(
    run_folder_id: str,
    image_id: str,
    carrier: str,
    drive_account: str,
) -> Path:
    safe_name = json.dumps(
        [IMAGE_CACHE_VERSION, run_folder_id, image_id, carrier, drive_account],
        ensure_ascii=True,
    )
    cache_key = hashlib.sha256(safe_name.encode("utf-8")).hexdigest()
    return IMAGE_CACHE_DIR / f"{cache_key}.bin"


def clear_persisted_cache() -> None:
    if RUN_DATA_CACHE_PATH.exists():
        RUN_DATA_CACHE_PATH.unlink()
    if IMAGE_CACHE_DIR.exists():
        shutil.rmtree(IMAGE_CACHE_DIR)


def _run_stub(run: RunFolder) -> tuple[str, str, str, str, str, str, str, str, str, str]:
    return (
        run.folder_id,
        run.folder_name,
        run.customer_code,
        run.run_date.isoformat(),
        run.carrier,
        run.run_id or "",
        str(run.metadata.get("source", "")),
        str(run.metadata.get("drive_account", DEFAULT_DRIVE_ACCOUNT)),
        str(run.metadata.get("carrier_folder_id", "")),
        "1" if bool(run.metadata.get("volatile_recent", False)) else "0",
    )


def _run_from_stub(stub: tuple[str, str, str, str, str, str, str, str, str, str]) -> RunFolder:
    (
        folder_id,
        folder_name,
        customer_code,
        run_date_iso,
        carrier,
        run_id,
        source,
        drive_account,
        carrier_folder_id,
        volatile_recent,
    ) = stub
    return RunFolder(
        folder_id=folder_id,
        folder_name=folder_name,
        customer_code=customer_code,
        run_date=date.fromisoformat(run_date_iso),
        carrier=carrier,
        run_id=run_id or None,
        metadata={
            "source": source,
            "drive_account": drive_account or DEFAULT_DRIVE_ACCOUNT,
            "carrier_folder_id": carrier_folder_id,
            "volatile_recent": volatile_recent == "1",
        },
    )


@st.cache_data(ttl=300, show_spinner=False)
def _load_run_details_payload(
    run_stubs: tuple[tuple[str, str, str, str, str, str, str, str, str, str], ...],
    _refresh_token: int,
) -> dict[str, object]:
    runs = [_run_from_stub(stub) for stub in run_stubs]
    hydrated_runs = hydrate_runs(runs)
    return _serialize_runs_payload(hydrated_runs, [])


def load_run_details(
    runs: list[RunFolder],
    refresh_token: int,
) -> list[RunFolder]:
    if not runs:
        return []
    payload = _load_run_details_payload(
        tuple(_run_stub(run) for run in runs),
        refresh_token,
    )
    hydrated_runs, _ = _deserialize_runs_payload(payload)
    return hydrated_runs


def _auto_rotation_degrees(image: Image.Image, carrier: str) -> int:
    if carrier not in USB_CAMERA_AUTO_ROTATE_CARRIERS:
        return 0

    normalized = ImageOps.exif_transpose(image)
    width, height = normalized.size

    if height >= width:
        return 0

    return 90


def _transform_image_bytes(image_data: bytes, carrier: str) -> bytes:
    with Image.open(BytesIO(image_data)) as image:
        normalized = ImageOps.exif_transpose(image)
        total_rotation = _auto_rotation_degrees(image, carrier) % 360
        rotated = normalized.rotate(total_rotation, expand=True) if total_rotation else normalized.copy()
        output = BytesIO()
        save_format = image.format or "PNG"

        if save_format == "JPEG" and rotated.mode in ("RGBA", "LA"):
            rotated = rotated.convert("RGB")

        rotated.save(output, format=save_format)
        return output.getvalue()


def _load_google_drive_run_index(
    drive_service: DriveService,
    root_folder_id: str,
    parse_errors: list[ParseError],
    archive_cutoff_date: date,
    maximum_date: date | None = None,
    drive_account: str = DEFAULT_DRIVE_ACCOUNT,
) -> list[RunFolder]:
    runs: list[RunFolder] = []
    root_folder = drive_service.get_file(root_folder_id)
    child_folders = drive_service.list_child_folders(root_folder_id)

    direct_run_folders: list[dict] = []
    carrier_folders: list[dict] = []

    for child_folder in child_folders:
        try:
            parsed = parse_run_folder_name(child_folder["name"])
            if parsed.run_date >= archive_cutoff_date:
                direct_run_folders.append(child_folder)
        except ValueError:
            carrier_folders.append(child_folder)

    if direct_run_folders:
        runs.extend(
            _build_google_run_index_for_folders(
                run_folders=direct_run_folders,
                carrier_name=root_folder["name"],
                carrier_folder_id=root_folder_id,
                parse_errors=parse_errors,
                minimum_date=archive_cutoff_date,
                maximum_date=maximum_date,
                drive_account=drive_account,
            )
        )

    for carrier_folder in carrier_folders:
        runs.extend(
            _build_google_run_index_for_folders(
                run_folders=drive_service.list_child_folders(carrier_folder["id"]),
                carrier_name=carrier_folder["name"],
                carrier_folder_id=carrier_folder["id"],
                parse_errors=parse_errors,
                minimum_date=archive_cutoff_date,
                maximum_date=maximum_date,
                drive_account=drive_account,
            )
        )

    return runs


def _build_google_run_index_for_folders(
    run_folders: list[dict],
    carrier_name: str,
    carrier_folder_id: str,
    parse_errors: list[ParseError],
    minimum_date: date,
    maximum_date: date | None = None,
    drive_account: str = DEFAULT_DRIVE_ACCOUNT,
) -> list[RunFolder]:
    runs: list[RunFolder] = []

    for run_folder in run_folders:
        folder_name = run_folder["name"]
        try:
            parsed = parse_run_folder_name(folder_name)
        except ValueError as exc:
            parse_errors.append(
                ParseError(
                    folder_id=run_folder["id"],
                    folder_name=folder_name,
                    carrier=carrier_name,
                    reason=str(exc),
                )
            )
            continue

        if parsed.run_date < minimum_date:
            continue
        if maximum_date is not None and parsed.run_date > maximum_date:
            continue

        runs.append(
            RunFolder(
                folder_id=run_folder["id"],
                folder_name=folder_name,
                customer_code=parsed.customer_code,
                run_date=parsed.run_date,
                carrier=carrier_name,
                run_id=parsed.run_id,
                metadata={
                    "carrier_folder_id": carrier_folder_id,
                    "source": "google_drive",
                    "drive_account": drive_account,
                },
            )
        )

    return runs


def hydrate_runs(runs: list[RunFolder]) -> list[RunFolder]:
    hydrated_runs: list[RunFolder] = []
    drive_service: tuple[str, DriveService] | None = None

    for run in runs:
        source = str(run.metadata.get("source", "google_drive"))
        if source == "google_drive":
            drive_account = str(
                run.metadata.get("drive_account", DEFAULT_DRIVE_ACCOUNT)
            )
            if drive_service is None or drive_service[0] != drive_account:
                drive_service = (drive_account, get_drive_service(drive_account))
            images, (qr_info, qr_source) = drive_service[1].list_run_folder_details(
                run.folder_id
            )
        else:
            files = [path for path in Path(run.folder_id).iterdir() if path.is_file()]
            images = _list_local_image_files(files)
            qr_info, qr_source = _extract_local_qr_info(files)

        hydrated_runs.append(
            RunFolder(
                folder_id=run.folder_id,
                folder_name=run.folder_name,
                customer_code=run.customer_code,
                run_date=run.run_date,
                carrier=run.carrier,
                run_id=run.run_id,
                images=images,
                qr_info=qr_info,
                qr_source=qr_source,
                metadata=dict(run.metadata),
            )
        )

    return hydrated_runs


def _list_local_image_files(files: list[Path]) -> list[ImageFile]:
    images: list[ImageFile] = []
    for file_path in files:
        mime_type, _ = mimetypes.guess_type(file_path.name)
        suffix = file_path.suffix.lower()
        if (mime_type or "").startswith("image/") or suffix in {
            ".jpg",
            ".jpeg",
            ".png",
            ".gif",
            ".bmp",
            ".webp",
            ".tif",
            ".tiff",
        }:
            images.append(
                ImageFile(
                    id=str(file_path),
                    name=file_path.name,
                    mime_type=mime_type or "application/octet-stream",
                    size=file_path.stat().st_size,
                )
            )
    return images


def _extract_local_qr_info(files: list[Path]) -> tuple[str, str | None]:
    for file_path in files:
        lower_name = file_path.name.lower()
        if lower_name in {"qr.txt", "qr.json"}:
            content = file_path.read_text(encoding="utf-8", errors="replace")
            if lower_name.endswith(".json"):
                try:
                    parsed = json.loads(content)
                    return json.dumps(parsed, indent=2, ensure_ascii=True), file_path.name
                except json.JSONDecodeError:
                    return content.strip() or "No QR info found", file_path.name
            return content.strip() or "No QR info found", file_path.name

    for file_path in files:
        if "qr" in file_path.name.lower():
            return file_path.name, "filename"

    return "No QR info found", None


@st.cache_data(ttl=300, show_spinner=False)
def load_run_images(
    run_folder_id: str,
    carrier: str,
    image_ids: tuple[str, ...],
    source: str,
    drive_account: str,
    volatile_recent: bool,
    _refresh_token: int,
) -> dict[str, bytes]:
    def read_or_cache_image(image_id: str) -> bytes:
        if source == "google_drive":
            drive_service = get_drive_service(drive_account)
            if volatile_recent:
                return _transform_image_bytes(
                    drive_service.download_file_bytes(image_id),
                    carrier,
                )

            cache_path = _image_cache_path(
                run_folder_id,
                image_id,
                carrier,
                drive_account,
            )
            if cache_path.exists():
                return cache_path.read_bytes()

            image_data = _transform_image_bytes(
                drive_service.download_file_bytes(image_id),
                carrier,
            )
            IMAGE_CACHE_DIR.mkdir(parents=True, exist_ok=True)
            cache_path.write_bytes(image_data)
            return image_data

        return _transform_image_bytes(
            Path(image_id).read_bytes(),
            carrier,
        )

    return {
        image_id: read_or_cache_image(image_id)
        for image_id in image_ids
    }


def render_run_card(run: RunFolder, debug_mode: bool, refresh_token: int) -> None:
    has_qr_info = run.qr_info != "No QR info found"
    st.markdown(
        "\n".join(
            [item for item in [
                f"**Carrier:** {run.carrier}",
                f"**Run ID:** {run.run_id or 'N/A'}",
                f"**Folder:** `{run.folder_name}`",
                f"**QR Info:** `{run.qr_info}`" if has_qr_info and run.qr_source == "filename" else None,
                "**QR Info:**" if has_qr_info and run.qr_source != "filename" else None,
            ] if item]
        )
    )
    if has_qr_info and run.qr_source and run.qr_source != "filename":
        st.caption(f"Source: {run.qr_source}")
    if has_qr_info and run.qr_source != "filename":
        st.code(run.qr_info, language="json" if str(run.qr_source).endswith(".json") else "text")
    if debug_mode:
        st.caption(f"Folder ID: {run.folder_id}")
    image_bytes = load_run_images(
        run.folder_id,
        run.carrier,
        tuple(image.id for image in run.images),
        str(run.metadata.get("source", "google_drive")),
        str(run.metadata.get("drive_account", DEFAULT_DRIVE_ACCOUNT)),
        bool(run.metadata.get("volatile_recent", False)),
        refresh_token,
    )
    render_run_images(run, image_bytes, debug_mode)


def render_run_images(
    run: RunFolder,
    image_bytes: dict[str, bytes],
    debug_mode: bool,
) -> None:
    if not run.images:
        st.caption("No image files found in this run folder.")
        return

    gallery_items: list[dict[str, str]] = []
    for image in run.images:
        try:
            image_data = image_bytes[image.id]
            encoded_image = base64.b64encode(image_data).decode("ascii")
            mime_type = image.mime_type or "image/jpeg"
            gallery_items.append(
                {
                    "name": image.name,
                    "src": f"data:{mime_type};base64,{encoded_image}",
                }
            )
        except Exception as exc:  # pragma: no cover - Streamlit rendering fallback
            st.caption(f"{image.name} could not be rendered: {exc}")

    if not gallery_items:
        return

    gallery_key = hashlib.sha1(run.folder_id.encode("utf-8")).hexdigest()[:12]
    column_count = min(3, max(1, len(gallery_items)))
    row_count = ceil(len(gallery_items) / column_count)
    component_height = (row_count * 270) + 100
    gallery_json = json.dumps(gallery_items)
    debug_note = (
        "<div class='gallery-debug-note'>Auto rotation enabled</div>"
        if debug_mode
        else ""
    )

    components.html(
        f"""
        <div id="gallery-{gallery_key}" class="run-gallery-wrap">
          <div class="run-gallery">
            {"".join(
                f'''
                <button
                  type="button"
                  class="run-photo-card"
                  data-index="{index}"
                  aria-label="Open {html.escape(item["name"])}"
                >
                  <img src="" alt="{html.escape(item["name"])}" />
                  <div class="run-photo-caption">{html.escape(item["name"])}</div>
                </button>
                '''
                for index, item in enumerate(gallery_items)
            )}
          </div>
          {debug_note}
          <div class="run-lightbox" hidden>
            <button type="button" class="run-lightbox-close" aria-label="Close photo viewer">×</button>
            <button type="button" class="run-lightbox-nav run-lightbox-prev" aria-label="Previous photo">‹</button>
            <figure class="run-lightbox-figure">
              <img class="run-lightbox-image" src="" alt="" />
              <figcaption class="run-lightbox-caption"></figcaption>
            </figure>
            <button type="button" class="run-lightbox-nav run-lightbox-next" aria-label="Next photo">›</button>
          </div>
        </div>
        <style>
          body {{
            margin: 0;
          }}
          #gallery-{gallery_key} {{
            font-family: "Source Sans Pro", sans-serif;
          }}
          #gallery-{gallery_key} .run-gallery {{
            display: grid;
            grid-template-columns: repeat({column_count}, minmax(0, 1fr));
            gap: 12px;
          }}
          #gallery-{gallery_key} .run-photo-card {{
            border: 0;
            background: transparent;
            padding: 0;
            cursor: pointer;
            text-align: left;
          }}
          #gallery-{gallery_key} .run-photo-card img {{
            width: auto;
            max-width: 100%;
            height: 180px;
            object-fit: contain;
            background: #f8fafc;
            border-radius: 12px;
            border: 1px solid #d6dde8;
            display: block;
            margin: 0 auto;
          }}
          #gallery-{gallery_key} .run-photo-caption,
          #gallery-{gallery_key} .gallery-debug-note {{
            margin-top: 0.35rem;
            color: #516074;
            font-size: 0.9rem;
            line-height: 1.3;
          }}
          #gallery-{gallery_key} .run-lightbox {{
            position: fixed;
            inset: 0;
            background: rgba(15, 23, 42, 0.9);
            z-index: 9999;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 24px;
          }}
          #gallery-{gallery_key} .run-lightbox[hidden] {{
            display: none;
          }}
          #gallery-{gallery_key} .run-lightbox-figure {{
            margin: 0;
            max-width: min(1100px, calc(100vw - 180px));
            max-height: calc(100vh - 64px);
            display: flex;
            flex-direction: column;
            align-items: center;
            gap: 12px;
          }}
          #gallery-{gallery_key} .run-lightbox-image {{
            max-width: 100%;
            max-height: calc(100vh - 140px);
            object-fit: contain;
            border-radius: 14px;
          }}
          #gallery-{gallery_key} .run-lightbox-caption {{
            color: #ffffff;
            font-size: 1rem;
            text-align: center;
          }}
          #gallery-{gallery_key} .run-lightbox-nav,
          #gallery-{gallery_key} .run-lightbox-close {{
            position: absolute;
            border: 0;
            color: #ffffff;
            background: rgba(255, 255, 255, 0.16);
            cursor: pointer;
            border-radius: 999px;
          }}
          #gallery-{gallery_key} .run-lightbox-nav {{
            top: 50%;
            transform: translateY(-50%);
            width: 52px;
            height: 52px;
            font-size: 2.4rem;
            line-height: 1;
          }}
          #gallery-{gallery_key} .run-lightbox-prev {{
            left: 24px;
          }}
          #gallery-{gallery_key} .run-lightbox-next {{
            right: 24px;
          }}
          #gallery-{gallery_key} .run-lightbox-close {{
            top: 20px;
            right: 24px;
            width: 44px;
            height: 44px;
            font-size: 2rem;
            line-height: 1;
          }}
        </style>
        <script>
          (() => {{
            const galleryRoot = document.getElementById("gallery-{gallery_key}");
            const photos = {gallery_json};
            const cards = Array.from(galleryRoot.querySelectorAll(".run-photo-card"));
            const thumbs = Array.from(galleryRoot.querySelectorAll(".run-photo-card img"));
            const lightbox = galleryRoot.querySelector(".run-lightbox");
            const lightboxImage = galleryRoot.querySelector(".run-lightbox-image");
            const lightboxCaption = galleryRoot.querySelector(".run-lightbox-caption");
            const previousButton = galleryRoot.querySelector(".run-lightbox-prev");
            const nextButton = galleryRoot.querySelector(".run-lightbox-next");
            const closeButton = galleryRoot.querySelector(".run-lightbox-close");
            const hostFrame = window.frameElement;
            const originalFrameStyle = hostFrame
              ? {{
                  position: hostFrame.style.position,
                  inset: hostFrame.style.inset,
                  width: hostFrame.style.width,
                  height: hostFrame.style.height,
                  zIndex: hostFrame.style.zIndex,
                  border: hostFrame.style.border,
                }}
              : null;
            let activeIndex = 0;

            const renderPhoto = (index) => {{
              activeIndex = (index + photos.length) % photos.length;
              lightboxImage.src = photos[activeIndex].src;
              lightboxImage.alt = photos[activeIndex].name;
              lightboxCaption.textContent = `${{activeIndex + 1}} / ${{photos.length}} - ${{photos[activeIndex].name}}`;
            }};

            thumbs.forEach((thumb, index) => {{
              thumb.src = photos[index].src;
            }});

            const openLightbox = (index) => {{
              renderPhoto(index);
              if (hostFrame) {{
                hostFrame.style.position = "fixed";
                hostFrame.style.inset = "0";
                hostFrame.style.width = "100vw";
                hostFrame.style.height = "100vh";
                hostFrame.style.zIndex = "999999";
                hostFrame.style.border = "0";
              }}
              lightbox.hidden = false;
              closeButton.focus();
            }};

            const closeLightbox = () => {{
              lightbox.hidden = true;
              lightboxImage.src = "";
              if (hostFrame && originalFrameStyle) {{
                hostFrame.style.position = originalFrameStyle.position;
                hostFrame.style.inset = originalFrameStyle.inset;
                hostFrame.style.width = originalFrameStyle.width;
                hostFrame.style.height = originalFrameStyle.height;
                hostFrame.style.zIndex = originalFrameStyle.zIndex;
                hostFrame.style.border = originalFrameStyle.border;
              }}
            }};

            cards.forEach((card) => {{
              card.addEventListener("click", () => {{
                openLightbox(Number(card.dataset.index || 0));
              }});
            }});

            previousButton.addEventListener("click", () => renderPhoto(activeIndex - 1));
            nextButton.addEventListener("click", () => renderPhoto(activeIndex + 1));
            closeButton.addEventListener("click", closeLightbox);
            lightbox.addEventListener("click", (event) => {{
              if (event.target === lightbox) {{
                closeLightbox();
              }}
            }});

            document.addEventListener("keydown", (event) => {{
              if (lightbox.hidden) {{
                return;
              }}
              if (event.key === "ArrowLeft") {{
                event.preventDefault();
                renderPhoto(activeIndex - 1);
              }}
              if (event.key === "ArrowRight") {{
                event.preventDefault();
                renderPhoto(activeIndex + 1);
              }}
              if (event.key === "Escape") {{
                event.preventDefault();
                closeLightbox();
              }}
            }});
          }})();
        </script>
        """,
        height=component_height,
        scrolling=False,
    )


def toggle_customer(customer_code: str) -> None:
    expanded_customers = set(st.session_state.get("expanded_customers", []))
    if customer_code in expanded_customers:
        expanded_customers.remove(customer_code)
    else:
        expanded_customers.add(customer_code)
    st.session_state["expanded_customers"] = sorted(expanded_customers)


def render_customer_header(customer_code: str, customer_runs: list[RunFolder], is_expanded: bool) -> None:
    total_images = sum(len(run.images) for run in customer_runs)
    carriers = sorted({run.carrier for run in customer_runs})
    card_class = "customer-card customer-card-open" if is_expanded else "customer-card"
    st.markdown(
        (
            f"<div class='{card_class}'>"
            f"<div class='customer-card-title'>{customer_code}</div>"
            f"<div class='customer-card-meta'>"
            f"{len(customer_runs)} runs | {total_images} images | "
            f"Carriers: {', '.join(carriers)}"
            f"</div>"
            f"</div>"
        ),
        unsafe_allow_html=True,
    )


def _hal_location_prefix(location: str | None) -> str:
    if not location:
        return ""
    return location[:2]


def _hal_customer_prefix(customer_code: str | None) -> str:
    if not customer_code:
        return ""
    if customer_code[:1].isdigit():
        return customer_code[:3]
    return customer_code[:2]


def _strip_hal_leading_g(location: str) -> str:
    if location[:1].lower() == "g":
        return location[1:].lstrip()
    return location


def _select_hal_sheet_name(sheet_names: list[str]) -> str:
    if "ERP_PASTE" in sheet_names:
        return "ERP_PASTE"
    if "Blad1" in sheet_names:
        return "Blad1"
    return sheet_names[0]


def _load_hal_rows(upload_name: str, upload_bytes: bytes) -> list[list[object]]:
    suffix = Path(upload_name).suffix.lower()

    if suffix == ".xls":
        workbook = xlrd.open_workbook(file_contents=upload_bytes)
        sheet_name = _select_hal_sheet_name(workbook.sheet_names())
        sheet = workbook.sheet_by_name(sheet_name)
        return [sheet.row_values(index) for index in range(sheet.nrows)]

    workbook = load_workbook(filename=BytesIO(upload_bytes), data_only=True, read_only=True)
    sheet_name = _select_hal_sheet_name(workbook.sheetnames)
    worksheet = workbook[sheet_name]
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def _parse_halindeling(upload_name: str, upload_bytes: bytes) -> list[dict[str, str]]:
    rows = _load_hal_rows(upload_name, upload_bytes)
    parsed: list[dict[str, str]] = []
    current_location: str | None = None

    for row in rows:
        location_value = row[0] if len(row) > 0 else None
        customer_value = row[1] if len(row) > 1 else None
        is_header = False

        if isinstance(location_value, str):
            location_text = location_value.strip()
            if (
                not location_text
                or location_text.startswith("Hal:")
                or location_text.startswith("---")
                or location_text.startswith("#")
                or location_text == "Locatie"
            ):
                is_header = True
            else:
                current_location = location_text

        if is_header:
            continue

        if isinstance(customer_value, str) and customer_value.strip() and current_location:
            parsed.append({"location": current_location, "customer": customer_value.strip()})

    return parsed


def _build_hal_dataset(upload_name: str, upload_bytes: bytes) -> dict[str, object]:
    data = _parse_halindeling(upload_name, upload_bytes)
    if not data:
        raise ValueError("Geen geldige halindeling-data gevonden")

    loc_prefixes = sorted({_hal_location_prefix(item["location"]) for item in data if _hal_location_prefix(item["location"])})
    cust_prefixes = sorted({_hal_customer_prefix(item["customer"]) for item in data if _hal_customer_prefix(item["customer"])})

    cust_by_loc: dict[str, list[str]] = {}
    grouped: dict[str, set[str]] = defaultdict(set)
    for item in data:
        grouped[_hal_location_prefix(item["location"])].add(_hal_customer_prefix(item["customer"]))
    for loc_prefix, prefixes in grouped.items():
        cust_by_loc[loc_prefix] = sorted(prefix for prefix in prefixes if prefix)

    return {
        "file_name": upload_name,
        "rows": data,
        "loc_prefixes": loc_prefixes,
        "cust_prefixes": cust_prefixes,
        "visible_cust_prefixes": list(cust_prefixes),
        "cust_by_loc": cust_by_loc,
    }


def _set_hal_checkbox_group(group_key: str, prefixes: list[str], checked: bool) -> None:
    st.session_state[group_key] = list(prefixes) if checked else []
    for prefix in prefixes:
        st.session_state[f"{group_key}_{prefix}"] = checked


def _render_hal_checkbox_grid(prefixes: list[str], group_key: str, columns_count: int = 6) -> list[str]:
    selected = set(st.session_state.get(group_key, []))
    columns = st.columns(columns_count)
    chosen: list[str] = []

    for index, prefix in enumerate(prefixes):
        checkbox_key = f"{group_key}_{prefix}"
        if checkbox_key not in st.session_state:
            st.session_state[checkbox_key] = prefix in selected
        with columns[index % columns_count]:
            checked = st.checkbox(prefix, key=checkbox_key)
        if checked:
            chosen.append(prefix)

    st.session_state[group_key] = chosen
    return chosen


def _generate_hal_pdf_bytes(rows: list[dict[str, str]], chosen_locations: list[str], chosen_customers: list[str]) -> bytes:
    filtered = []
    for item in rows:
        loc_prefix = _hal_location_prefix(item["location"])
        cust_prefix = _hal_customer_prefix(item["customer"])
        loc_ok = not chosen_locations or loc_prefix in chosen_locations
        cust_ok = not chosen_customers or cust_prefix in chosen_customers
        if loc_ok and cust_ok:
            filtered.append(item)

    unique_rows: list[dict[str, str]] = []
    seen_customers: set[str] = set()
    for item in filtered:
        customer = item["customer"]
        if customer in seen_customers:
            continue
        seen_customers.add(customer)
        unique_rows.append(item)

    if not unique_rows:
        raise ValueError("Geen klanten gevonden voor deze filters")

    page_width = 10 * (72 / 2.54)
    page_height = 15 * (72 / 2.54)
    margin = 0.4 * (72 / 2.54)
    gap = 0.4 * (72 / 2.54)
    location_ratio = 4

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=(page_width, page_height))

    def fit_font_size(text: str, max_width: float, max_height: float) -> int:
        size = 1
        while size < 600:
            next_size = size + 1
            text_width = stringWidth(text, "Helvetica-Bold", next_size)
            text_height = next_size
            if text_width > max_width * 0.97 or text_height > max_height * 0.97:
                break
            size = next_size
        return size

    for item in unique_rows:
        location = _strip_hal_leading_g(item["location"])
        customer = item["customer"]

        pdf.saveState()
        pdf.translate(0, page_height)
        pdf.rotate(-90)

        width = page_height
        height = page_width
        inner_width = width - (2 * margin)
        available_height = height - (2 * margin) - gap
        customer_height = available_height / (location_ratio + 1)
        location_height = (available_height * location_ratio) / (location_ratio + 1)

        customer_size = fit_font_size(customer, inner_width, customer_height)
        customer_width = stringWidth(customer, "Helvetica-Bold", customer_size)
        customer_text_height = customer_size * 0.72
        customer_y = height - margin - customer_height + ((customer_height - customer_text_height) / 2) - (customer_size * 0.1)
        pdf.setFont("Helvetica-Bold", customer_size)
        pdf.drawString((width - customer_width) / 2, customer_y, customer)

        target_location_size = customer_size * location_ratio
        max_location_size = fit_font_size(location, inner_width, location_height)
        location_size = min(target_location_size, max_location_size)
        location_width = stringWidth(location, "Helvetica-Bold", location_size)
        location_text_height = location_size * 0.72
        location_y = margin + ((location_height - location_text_height) / 2) - (location_size * 0.1)
        pdf.setFont("Helvetica-Bold", location_size)
        pdf.drawString((width - location_width) / 2, location_y, location)

        pdf.restoreState()
        pdf.showPage()

    pdf.save()
    return buffer.getvalue()


def render_hal_locations_page() -> None:
    st.title("Hal Locations")
    st.caption("Upload een halindeling en download stickers met exact dezelfde filters en PDF-logica als de StickerPrinter app.")

    upload_col, action_col = st.columns([3, 1])
    with upload_col:
        upload = st.file_uploader(
            "1. Upload de halindeling",
            type=["xlsx", "xls"],
            key="hal_locations_upload",
            help="Selecteer het Halindeling .xlsx of .xls bestand.",
        )
    with action_col:
        st.write("")
        st.write("")
        upload_clicked = st.button("Upload", use_container_width=True, key="hal_locations_upload_btn")

    if upload_clicked:
        if upload is None:
            st.error("Selecteer eerst een bestand")
        else:
            try:
                dataset = _build_hal_dataset(upload.name, upload.getvalue())
            except Exception as exc:
                st.error(f"Fout: {exc}")
            else:
                st.session_state["hal_locations_dataset"] = dataset
                _set_hal_checkbox_group("hal_locations_selected_loc", dataset["loc_prefixes"], False)
                _set_hal_checkbox_group("hal_locations_selected_cust", dataset["cust_prefixes"], False)
                st.success(
                    f"OK - {len(dataset['rows'])} regels, {len(dataset['loc_prefixes'])} locatie-prefixen, {len(dataset['cust_prefixes'])} klant-prefixen"
                )

    dataset = st.session_state.get("hal_locations_dataset")
    if not dataset:
        return

    st.markdown("### 2. Selecteer locaties")
    st.caption("Eerste 2 tekens van de locatiecode. Bijvoorbeeld `gK`, `gL`, `bA`, `eT`.")
    loc_action_all, loc_action_none = st.columns([1, 1])
    with loc_action_all:
        if st.button("Alles aanvinken", key="hal_loc_all", use_container_width=True):
            _set_hal_checkbox_group("hal_locations_selected_loc", dataset["loc_prefixes"], True)
    with loc_action_none:
        if st.button("Niets aanvinken", key="hal_loc_none", use_container_width=True):
            _set_hal_checkbox_group("hal_locations_selected_loc", dataset["loc_prefixes"], False)
    chosen_locations = _render_hal_checkbox_grid(dataset["loc_prefixes"], "hal_locations_selected_loc")

    st.markdown("### 3. Selecteer klantcodes (optioneel)")
    st.caption(
        "Filter op klantcode-prefix. Codes met cijfer aan het begin gebruiken 3 tekens, codes met letter aan het begin 2 tekens. Leeg laten betekent alle klanten op de gekozen locaties."
    )
    cust_prefixes = list(dataset.get("visible_cust_prefixes", dataset["cust_prefixes"]))
    cust_action_all, cust_action_none, cust_action_filter = st.columns([1, 1, 2])
    with cust_action_all:
        if st.button("Alles aanvinken", key="hal_cust_all", use_container_width=True):
            _set_hal_checkbox_group("hal_locations_selected_cust", cust_prefixes, True)
    with cust_action_none:
        if st.button("Niets aanvinken", key="hal_cust_none", use_container_width=True):
            _set_hal_checkbox_group("hal_locations_selected_cust", cust_prefixes, False)
    with cust_action_filter:
        if st.button("Alleen prefixen op gekozen locaties", key="hal_cust_filter", use_container_width=True):
            if not chosen_locations:
                st.warning("Vink eerst locaties aan")
            else:
                allowed_prefixes = sorted({
                    prefix
                    for location_prefix in chosen_locations
                    for prefix in dataset["cust_by_loc"].get(location_prefix, [])
                })
                dataset["visible_cust_prefixes"] = allowed_prefixes
                st.session_state["hal_locations_dataset"] = dataset
                _set_hal_checkbox_group("hal_locations_selected_cust", dataset["cust_prefixes"], False)
                _set_hal_checkbox_group("hal_locations_selected_cust", allowed_prefixes, False)
                cust_prefixes = allowed_prefixes
    chosen_customers = _render_hal_checkbox_grid(cust_prefixes, "hal_locations_selected_cust")

    st.markdown("### 4. Genereer PDF")
    if not chosen_locations:
        st.info("Vink minstens een locatie aan om de stickers te genereren.")
        st.download_button(
            "Download stickers PDF",
            data=b"",
            file_name="stickers.pdf",
            disabled=True,
            use_container_width=True,
        )
        return

    try:
        pdf_bytes = _generate_hal_pdf_bytes(dataset["rows"], chosen_locations, chosen_customers)
    except ValueError as exc:
        st.error(str(exc))
        st.download_button(
            "Download stickers PDF",
            data=b"",
            file_name="stickers.pdf",
            disabled=True,
            use_container_width=True,
        )
        return

    st.download_button(
        "Download stickers PDF",
        data=pdf_bytes,
        file_name=f"stickers_{date.today().isoformat()}.pdf",
        mime="application/pdf",
        use_container_width=True,
    )


def render_photo_dashboard() -> None:
    st.title("Sjaak vd Vijver Expedition Photo Dashboard")
    st.caption("Choose departure date and optionally filter by customer code")

    sync_status = _read_index_sync_status()
    sync_signature = f"{sync_status.get('state','')}|{sync_status.get('updated_at','')}"
    previous_sync_signature = st.session_state.get("index_sync_signature")
    if sync_signature != previous_sync_signature:
        st.session_state["index_sync_signature"] = sync_signature
        if sync_status.get("state") == "completed":
            _load_run_data_payload.clear()
            _load_run_details_payload.clear()
            load_run_images.clear()

    st.sidebar.header("Controls")
    debug_mode = st.sidebar.checkbox("Debug mode", value=False)
    sync_running = sync_status.get("state") == "running"
    rebuild_index_requested = st.sidebar.button(
        "Rebuild run index",
        use_container_width=True,
        disabled=sync_running,
    )
    clear_cache_requested = st.sidebar.button("Clear all caches", use_container_width=True)

    refresh_token = st.session_state.get("refresh_token", 0)
    if rebuild_index_requested:
        if _start_index_sync("rebuild"):
            st.sidebar.caption("Background rebuild started. You can keep using the dashboard while it runs.")
        else:
            st.sidebar.error("Could not start the background rebuild script.")
    elif clear_cache_requested:
        refresh_token += 1
        st.session_state["refresh_token"] = refresh_token
        clear_persisted_cache()
        _load_run_data_payload.clear()
        _load_run_details_payload.clear()
        load_run_images.clear()
        st.sidebar.caption("All saved caches cleared. Data will be reloaded from source.")
    elif RUN_DATA_CACHE_PATH.exists():
        cache_timestamp = _read_persisted_run_cache_timestamp()
        cache_label = (
            cache_timestamp.astimezone().strftime("%Y-%m-%d %H:%M:%S")
            if cache_timestamp is not None
            else "unknown time"
        )
        st.sidebar.caption(
            f"Using saved run index from {cache_label}. Use 'Refresh this date' for new photos on one date, or 'Rebuild run index' for a full rescan."
        )
    else:
        st.sidebar.caption("No saved run index yet. The next load will scan the source folders and store it locally.")

    sync_mode = str(sync_status.get("mode", "")).replace("_", " ").strip()
    sync_updated_label = _format_timestamp(sync_status.get("updated_at"))
    sync_started_label = _format_timestamp(sync_status.get("started_at"))
    if sync_running:
        detail = f"Running {sync_mode}" if sync_mode else "Background sync running"
        if sync_started_label:
            detail = f"{detail} since {sync_started_label}"
        st.sidebar.info(detail)
    elif sync_status.get("state") == "completed":
        detail = "Background sync completed"
        if sync_updated_label:
            detail = f"{detail} at {sync_updated_label}"
        st.sidebar.success(detail)
    elif sync_status.get("state") == "failed":
        detail = "Background sync failed"
        if sync_updated_label:
            detail = f"{detail} at {sync_updated_label}"
        st.sidebar.error(detail)
        error_message = sync_status.get("error")
        if isinstance(error_message, str) and error_message.strip():
            st.sidebar.caption(error_message)

    _render_background_sync_autorefresh(sync_running)

    try:
        with st.spinner("Scanning Google Drive and local archive folders..."):
            runs, parse_errors = load_run_data(refresh_token)
    except (RuntimeError, HttpError, OSError) as exc:
        st.error(f"Unable to load data from Google Drive: {exc}")
        st.stop()

    render_parse_errors(parse_errors, debug_mode)

    available_dates = sorted({run.run_date for run in runs})
    if not available_dates:
        st.warning("No valid run folders were found under the configured root folder.")
        return

    default_date = available_dates[-1]
    filter_col, refresh_col, search_col, metric_customer_col, metric_run_col, metric_image_col = st.columns(
        [2.2, 1.4, 2.0, 1, 1, 1]
    )
    with filter_col:
        selected_date = st.date_input(
            "Filter by date",
            value=default_date,
            min_value=available_dates[0],
            max_value=available_dates[-1],
            key="selected_date",
        )
    with refresh_col:
        st.write("")
        refresh_selected_date = st.button(
            "Refresh this date",
            use_container_width=True,
            disabled=sync_running,
        )
    if refresh_selected_date:
        if _start_index_sync("refresh_date", selected_date):
            st.success(f"Background refresh started for {selected_date.isoformat()}.")
        else:
            st.error("Could not start the background refresh script.")
    with search_col:
        search_term = st.text_input("Search customer code", placeholder="cust123")

    filtered_runs = [run for run in runs if run.run_date == selected_date]
    if search_term.strip():
        search_lower = search_term.strip().lower()
        filtered_runs = [
            run for run in filtered_runs if search_lower in run.customer_code.lower()
        ]
    filtered_runs = load_run_details(filtered_runs, refresh_token)

    customer_count, run_count, image_count = (
        len({run.customer_code for run in filtered_runs}),
        len(filtered_runs),
        sum(len(run.images) for run in filtered_runs),
    )
    metric_customer_col.metric("Customers", customer_count)
    metric_run_col.metric("Runs", run_count)
    metric_image_col.metric("Images", image_count)

    grouped_runs = group_runs_by_customer(filtered_runs)
    if not grouped_runs:
        st.info("No runs match the selected filters.")
        return

    for customer_code, customer_runs in grouped_runs.items():
        expanded_customers = set(st.session_state.get("expanded_customers", []))
        is_expanded = customer_code in expanded_customers
        header_col, action_col = st.columns([6, 1])
        with header_col:
            render_customer_header(customer_code, customer_runs, is_expanded)
        with action_col:
            st.write("")
            st.write("")
            st.button(
                "Collapse" if is_expanded else "Expand",
                key=f"toggle-{customer_code}",
                on_click=toggle_customer,
                args=(customer_code,),
                use_container_width=True,
            )
        if is_expanded:
            for run in customer_runs:
                with st.container():
                    st.markdown("<div class='run-shell'>", unsafe_allow_html=True)
                    st.subheader(f"{run.carrier} | {run.run_id or 'No run ID'}")
                    render_run_card(run, debug_mode, refresh_token)
                    st.markdown("</div>", unsafe_allow_html=True)

    if debug_mode:
        st.sidebar.subheader("Debug summary")
        st.sidebar.write(f"Root folder ID: `{get_setting('GOOGLE_DRIVE_ROOT_FOLDER_ID') or ''}`")
        st.sidebar.write(f"Local archive root: `{get_setting('LOCAL_ARCHIVE_ROOT') or ''}`")
        carriers_by_date: dict[date, set[str]] = defaultdict(set)
        for run in runs:
            carriers_by_date[run.run_date].add(run.carrier)
        st.sidebar.write(f"Dates found: {len(available_dates)}")
        st.sidebar.write(
            f"Carriers on selected date: {len(carriers_by_date.get(selected_date, set()))}"
        )


def main() -> None:
    st.sidebar.header("Menu")
    selected_page = st.sidebar.radio(
        "Ga naar",
        ["Photo Dashboard", "Hal Locations"],
        key="main_menu",
    )

    if selected_page == "Hal Locations":
        render_hal_locations_page()
        return

    render_photo_dashboard()


if __name__ == "__main__":
    main()
