#!/usr/bin/env python3
import base64
import io
import json
import re
import sys
from copy import copy
from collections import defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile
import xml.etree.ElementTree as ET

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"

CATEGORY_DEFINITIONS = {
    "508": {"label": "Flowers", "slug": "bloemen"},
    "515": {"label": "Accessories", "slug": "acc"},
    "1000": {"label": "Bouquets / BQ", "slug": "BQ"},
    "920": {"label": "Plants", "slug": "PLANTEN"},
}
CATEGORY_ORDER = ["515", "1000", "508", "920"]
EXPECTED_COLUMNS = [
    "itemIdClientSystem", "itemNumber", "materialNumber", "invoiceIdClientSystem", "grossMassValue", "grossMassUnit",
    "netMassValue", "netMassUnit", "netPriceValue", "netPriceCurrencyIso", "value", "originCountryCode",
    "preferentialOriginCountryCode", "classificationType", "classificationValue", "goodsDescriptionText", "quantityValue",
    "quantityUnit", "packages", "order", "packageCode", "taricCode", "fullClassificationCode", "vbnCode", "vbnDescription",
]
ZERO = Decimal("0")
DEFAULT_VALUE_TOLERANCE = Decimal("0.01")
DEFAULT_WEIGHT_TOLERANCE = Decimal("0.001")
DEFAULT_QUANTITY_TOLERANCE = Decimal("0")
DEFAULT_PACKAGES_TOLERANCE = Decimal("0")
BLOCKING_WARNING_CODES = {
    "missing_required_columns",
    "missing_commodity_code",
    "missing_origin",
    "missing_value",
    "missing_quantity",
    "missing_packages",
    "unsupported_weight_unit",
    "negative_value",
    "blank_classification_value",
    "duplicate_invoice_number",
}


def dec(value, default=ZERO):
    if isinstance(value, Decimal):
        return value
    text = str(value or "").strip()
    if not text:
        return default
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return default


def strip_invalid_xml_chars(value):
    text = str(value or "")
    return "".join(
        ch for ch in text
        if ch in ("\t", "\n", "\r") or ord(ch) >= 32
    )


def clean_text(value):
    return strip_invalid_xml_chars(value).strip()


def lower_key(value):
    return clean_text(value).lower()


def json_decimal(value):
    if isinstance(value, Decimal):
        s = format(value, "f")
        if "." in s:
            s = s.rstrip("0").rstrip(".") or "0"
        return s
    if isinstance(value, dict):
        return {key: json_decimal(item) for key, item in value.items()}
    if isinstance(value, list):
        return [json_decimal(item) for item in value]
    return value


def col_letters_to_index(ref):
    letters = "".join(ch for ch in ref if ch.isalpha())
    total = 0
    for ch in letters:
        total = (total * 26) + (ord(ch.upper()) - 64)
    return total - 1


def escape_xml(value):
    text = strip_invalid_xml_chars(value)
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


def parse_xlsx_rows(raw_bytes):
    with ZipFile(io.BytesIO(raw_bytes)) as zf:
        shared = []
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall(f"{{{NS_MAIN}}}si"):
                parts = []
                for t in si.iterfind(f".//{{{NS_MAIN}}}t"):
                    parts.append(t.text or "")
                shared.append("".join(parts))
        workbook = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
        first_sheet = workbook.find(f"{{{NS_MAIN}}}sheets")[0]
        target = rel_map[first_sheet.attrib[f"{{{NS_REL}}}id"]]
        worksheet = ET.fromstring(zf.read(f"xl/{target}"))
        sheet_data = worksheet.find(f"{{{NS_MAIN}}}sheetData")
        rows = []
        for row in sheet_data.findall(f"{{{NS_MAIN}}}row"):
            values = []
            for cell in row.findall(f"{{{NS_MAIN}}}c"):
                idx = col_letters_to_index(cell.attrib.get("r", "A1"))
                while len(values) <= idx:
                    values.append("")
                cell_type = cell.attrib.get("t")
                if cell_type == "inlineStr":
                    text_nodes = cell.findall(f".//{{{NS_MAIN}}}t")
                    values[idx] = "".join(node.text or "" for node in text_nodes)
                    continue
                v = cell.find(f"{{{NS_MAIN}}}v")
                text = v.text if v is not None and v.text is not None else ""
                if cell_type == "s" and text:
                    try:
                        values[idx] = shared[int(text)]
                    except Exception:
                        values[idx] = text
                else:
                    values[idx] = text
            rows.append(values)
        return rows


def xlsx_rows_to_cell_map(rows):
    cell_map = {}
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            if value is None:
                continue
            text = str(value)
            if text == "":
                continue
            cell_map[(row_index, col_index)] = text
    return cell_map


def xlsx_cell(cells, row_index, col_index):
    return clean_text(cells.get((row_index, col_index), ""))


def parse_prefixed_value(value, prefix):
    text = clean_text(value)
    prefix_text = clean_text(prefix)
    if prefix_text and text.lower().startswith(prefix_text.lower()):
        return clean_text(text[len(prefix_text):])
    return text


def extract_company_settings_from_invoice_rows(rows):
    cells = xlsx_rows_to_cell_map(rows)
    return {
        "company_name": xlsx_cell(cells, 49, 3) or xlsx_cell(cells, 28, 3) or xlsx_cell(cells, 38, 3),
        "address": "\n".join(filter(None, [xlsx_cell(cells, 50, 3) or xlsx_cell(cells, 29, 3) or xlsx_cell(cells, 39, 3), xlsx_cell(cells, 51, 3) or xlsx_cell(cells, 30, 3) or xlsx_cell(cells, 40, 3)])),
        "phone": parse_prefixed_value(xlsx_cell(cells, 52, 3) or xlsx_cell(cells, 31, 3) or xlsx_cell(cells, 41, 3), "tel"),
        "email": parse_prefixed_value(xlsx_cell(cells, 53, 3) or xlsx_cell(cells, 32, 3) or xlsx_cell(cells, 42, 3), "email :"),
        "website": parse_prefixed_value(xlsx_cell(cells, 54, 3) or xlsx_cell(cells, 33, 3) or xlsx_cell(cells, 43, 3), "web :"),
        "vat_number": parse_prefixed_value(xlsx_cell(cells, 49, 7) or xlsx_cell(cells, 28, 7) or xlsx_cell(cells, 38, 7), "VAT nr :"),
        "eori_number": parse_prefixed_value(xlsx_cell(cells, 50, 7) or xlsx_cell(cells, 29, 7) or xlsx_cell(cells, 39, 7), "EORI nr:"),
        "chamber_of_commerce_number": parse_prefixed_value(xlsx_cell(cells, 51, 7) or xlsx_cell(cells, 30, 7) or xlsx_cell(cells, 40, 7), "Chamber of Commerce :"),
        "iban": parse_prefixed_value(xlsx_cell(cells, 52, 7) or xlsx_cell(cells, 31, 7) or xlsx_cell(cells, 41, 7), "IBAN :"),
        "bic_swift": parse_prefixed_value(xlsx_cell(cells, 53, 7) or xlsx_cell(cells, 32, 7) or xlsx_cell(cells, 42, 7), "BIC/SWIFT :"),
        "rex_registration": parse_prefixed_value(xlsx_cell(cells, 54, 7) or xlsx_cell(cells, 33, 7) or xlsx_cell(cells, 43, 7), "rex registration :"),
        "default_footer_text": xlsx_cell(cells, 55, 3) or xlsx_cell(cells, 34, 3) or xlsx_cell(cells, 44, 3),
        "preferential_origin_declaration": xlsx_cell(cells, 56, 3) or xlsx_cell(cells, 35, 3) or xlsx_cell(cells, 45, 3),
    }



def extract_customer_from_invoice_rows(rows):
    cells = xlsx_rows_to_cell_map(rows)
    return {
        "customer_name": xlsx_cell(cells, 4, 2),
        "customer_address": "\n".join(filter(None, [xlsx_cell(cells, 5, 2), xlsx_cell(cells, 6, 2)])),
        "vat_number": parse_prefixed_value(xlsx_cell(cells, 7, 2), "VAT NR"),
        "eori_number": parse_prefixed_value(xlsx_cell(cells, 8, 2), "EORI NR"),
        "importer_number": parse_prefixed_value(xlsx_cell(cells, 9, 2), ""),
        "default_delivery_terms": xlsx_cell(cells, 13, 3),
        "default_city": xlsx_cell(cells, 6, 2),
        "default_uk_arrival_port": "",
        "default_currency": "",
        "default_invoice_language_text": "",
        "default_document_references": xlsx_cell(cells, 9, 2),
    }



def extract_export_defaults_from_export_rows(rows):
    cells = xlsx_rows_to_cell_map(rows)
    return {
        "destination_country": xlsx_cell(cells, 2, 4) or "GB",
        "regulation": xlsx_cell(cells, 2, 3) or "Export",
        "border_transport_mode": xlsx_cell(cells, 2, 10) or "Road",
        "border_transport_nationality": xlsx_cell(cells, 2, 11) or "NL",
        "customs_office_of_exit": xlsx_cell(cells, 2, 14),
        "location": xlsx_cell(cells, 2, 7),
        "delivery_terms": xlsx_cell(cells, 2, 12),
        "delivery_terms_city": xlsx_cell(cells, 2, 13),
        "currency": xlsx_cell(cells, 12, 14) or "GBP",
        "freight_costs": xlsx_cell(cells, 14, 14),
        "insurance": xlsx_cell(cells, 15, 14),
        "importer_field": xlsx_cell(cells, 6, 14),
        "vessel_field": xlsx_cell(cells, 9, 14),
        "phyto_fields": xlsx_cell(cells, 11, 14),
        "kcb_fields": xlsx_cell(cells, 11, 14),
        "certificate_fields": xlsx_cell(cells, 4, 14),
    }



def import_example_payload(payload):
    invoice_file = (payload.get("invoice_example") or {})
    export_file = (payload.get("export_example") or {})
    invoice_rows = parse_xlsx_rows(base64.b64decode(clean_text(invoice_file.get("content_base64")))) if clean_text(invoice_file.get("content_base64")) else []
    export_rows = parse_xlsx_rows(base64.b64decode(clean_text(export_file.get("content_base64")))) if clean_text(export_file.get("content_base64")) else []
    customer = extract_customer_from_invoice_rows(invoice_rows) if invoice_rows else {}
    company = extract_company_settings_from_invoice_rows(invoice_rows) if invoice_rows else {}
    export_defaults = extract_export_defaults_from_export_rows(export_rows) if export_rows else {}
    warnings = []
    if not invoice_rows:
        warnings.append("No invoice example uploaded.")
    if not export_rows:
        warnings.append("No export example uploaded.")
    return {
        "customer": customer,
        "company_settings": company,
        "export_defaults": export_defaults,
        "warnings": warnings,
    }


def parse_invoice_numbers(payload):
    explicit = payload.get("invoice_numbers_by_category") or {}
    result = {code: clean_text(explicit.get(code)) for code in CATEGORY_DEFINITIONS}
    raw = clean_text(payload.get("invoice_numbers"))
    if raw:
        tokens = [token for token in re.split(r"[\s,;/]+", raw.replace("-", "/")) if token]
        for code, token in zip(CATEGORY_ORDER, tokens):
            if not result.get(code):
                result[code] = token
    return result


def build_column_alias_map(payload, category_code):
    aliases = ((payload.get("column_mappings") or {}).get(category_code) or {}).get("aliases") or {}
    mapping = {}
    for expected in EXPECTED_COLUMNS:
        candidates = [expected] + list(aliases.get(expected, []))
        mapping[expected] = [lower_key(candidate) for candidate in candidates if clean_text(candidate)]
    return mapping


def resolve_header_indexes(headers, alias_map):
    normalized_headers = {lower_key(value): index for index, value in enumerate(headers)}
    resolved = {}
    for expected, aliases in alias_map.items():
        resolved[expected] = next((normalized_headers[alias] for alias in aliases if alias in normalized_headers), None)
    return resolved


def row_value(row, indexes, key):
    idx = indexes.get(key)
    if idx is None or idx >= len(row):
        return ""
    return row[idx]


def convert_weight_kg(value, unit):
    unit_norm = lower_key(unit)
    amount = dec(value)
    if unit_norm in {"gr", "g", "gram", "grams"}:
        return amount / Decimal("1000"), None
    if unit_norm in {"kg", "kilogram", "kilograms"}:
        return amount, None
    return amount, f"Unsupported unit '{unit}'"


def sum_dicts(items, fields):
    totals = {field: ZERO for field in fields}
    for item in items:
        for field in fields:
            totals[field] += dec(item.get(field))
    return totals


def compare_values(left, right, tolerance):
    difference = left - right
    return {
        "left": left,
        "right": right,
        "difference": difference,
        "ok": abs(difference) <= tolerance,
    }


def aggregate_rows(records, key_fields):
    grouped = {}
    for record in records:
        key = tuple(record[field] for field in key_fields)
        if key not in grouped:
            grouped[key] = {
                **{field: record[field] for field in key_fields},
                "quantity": ZERO,
                "gross_kg": ZERO,
                "net_kg": ZERO,
                "packages": ZERO,
                "customs_value": ZERO,
            }
        grouped[key]["quantity"] += record["quantity"]
        grouped[key]["gross_kg"] += record["gross_kg"]
        grouped[key]["net_kg"] += record["net_kg"]
        grouped[key]["packages"] += record["packages"]
        grouped[key]["customs_value"] += record["customs_value"]
    return list(grouped.values())


def build_audit_checks(category_totals, invoice_totals, combined_source_totals, combined_export_totals, payload):
    export_defaults = payload.get("export_defaults") or {}
    value_tol = dec(export_defaults.get("value_tolerance"), DEFAULT_VALUE_TOLERANCE)
    weight_tol = dec(export_defaults.get("weight_tolerance"), DEFAULT_WEIGHT_TOLERANCE)
    quantity_tol = dec(export_defaults.get("quantity_tolerance"), DEFAULT_QUANTITY_TOLERANCE)
    packages_tol = dec(export_defaults.get("packages_tolerance"), DEFAULT_PACKAGES_TOLERANCE)
    checks = []
    summary_rows = []
    field_definitions = [
        ("quantity", quantity_tol),
        ("gross_kg", weight_tol),
        ("net_kg", weight_tol),
        ("packages", packages_tol),
        ("customs_value", value_tol),
    ]
    for category_code, source in category_totals.items():
        invoice = invoice_totals.get(category_code, {"quantity": ZERO, "gross_kg": ZERO, "net_kg": ZERO, "packages": ZERO, "customs_value": ZERO})
        for field, tolerance in field_definitions:
            comparison = compare_values(dec(source.get(field)), dec(invoice.get(field)), tolerance)
            checks.append({"scope": "category", "category": category_code, "field": field, **comparison})
            summary_rows.append({
                "scope": "category",
                "group_label": f'{category_code} {CATEGORY_DEFINITIONS.get(category_code, {}).get("label", "")}'.strip(),
                "field": field,
                "dump_value": dec(source.get(field)),
                "invoice_value": dec(invoice.get(field)),
                "export_value": "",
                "invoice_difference": comparison["difference"],
                "export_difference": "",
                "status": "MATCH" if comparison["ok"] else "FAIL",
            })
    for field, tolerance in field_definitions:
        comparison = compare_values(dec(combined_source_totals.get(field)), dec(combined_export_totals.get(field)), tolerance)
        checks.append({"scope": "combined_export", "field": field, **comparison})
        summary_rows.append({
            "scope": "combined_export",
            "group_label": "Combined export",
            "field": field,
            "dump_value": dec(combined_source_totals.get(field)),
            "invoice_value": "",
            "export_value": dec(combined_export_totals.get(field)),
            "invoice_difference": "",
            "export_difference": comparison["difference"],
            "status": "MATCH" if comparison["ok"] else "FAIL",
        })
    return checks, summary_rows


def rows_to_dense_map(rows):
    mapping = {}
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            if value is None or value == "":
                continue
            mapping[(row_idx, col_idx)] = value
    return mapping


def excel_date_serial(date_text):
    text = clean_text(date_text)
    if not text:
        return ""
    try:
        dt = datetime.strptime(text, "%Y-%m-%d")
    except ValueError:
        return text
    base = datetime(1899, 12, 30)
    return (dt - base).days


def normalize_invoice_hs_code(value):
    text = clean_text(value)
    if not text:
        return ""
    if text.isdigit() and len(text) >= 8:
        trimmed = text.lstrip("0") or "0"
        if len(trimmed) > 7:
            trimmed = trimmed[:-2] or trimmed
        return trimmed
    return text


def column_name(index):
    result = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        result = chr(65 + rem) + result
    return result


STYLE_NAME_TO_ID = {
    "default": 0,
    "bold": 1,
    "export_header": 2,
    "export_goods": 3,
    "detail_label": 4,
    "detail_value": 5,
    "table_header": 6,
    "footer": 7,
    "detail_value_plain": 8,
    "detail_label_plain": 9,
}


def png_dimensions(raw_bytes):
    if raw_bytes[:8] != b"\x89PNG\r\n\x1a\n":
        return 220, 70
    width = int.from_bytes(raw_bytes[16:20], "big")
    height = int.from_bytes(raw_bytes[20:24], "big")
    return width or 220, height or 70


def resolve_logo_image(company):
    candidates = []
    logo_name = clean_text((company or {}).get("logo_name"))
    if logo_name:
        candidates.append(logo_name.lstrip("/"))
    candidates.append("logosjaak.png")
    base_dir = Path(__file__).resolve().parent.parent / "public"
    for relative in candidates:
        file_path = base_dir / relative
        if file_path.exists() and file_path.is_file():
            raw = file_path.read_bytes()
            width, height = png_dimensions(raw)
            return {
                "bytes": raw,
                "ext": file_path.suffix.lower().lstrip(".") or "png",
                "width": width,
                "height": height,
                "from_col": 0,
                "from_row": 0,
            }
    return None


def build_sheet_xml(cell_map, style_map=None, column_widths=None, row_heights=None, include_drawing=False):
    style_map = style_map or {}
    column_widths = column_widths or {}
    row_heights = row_heights or {}
    if not cell_map:
        cell_map = {(1, 1): ""}
    max_row = max(max((row for row, _ in cell_map), default=1), max(row_heights.keys(), default=1))
    max_col = max(max((col for _, col in cell_map), default=1), max(column_widths.keys(), default=1))
    lines = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        f'<worksheet xmlns="{NS_MAIN}" xmlns:r="{NS_REL}">',
        f'<dimension ref="A1:{column_name(max_col)}{max_row}"/>',
        '<sheetViews><sheetView workbookViewId="0"/></sheetViews>',
    ]
    if column_widths:
        lines.append('<cols>')
        for col_index in sorted(column_widths):
            lines.append(f'<col min="{col_index}" max="{col_index}" width="{column_widths[col_index]}" customWidth="1"/>')
        lines.append('</cols>')
    lines.append('<sheetFormatPr defaultRowHeight="15"/>')
    lines.append('<sheetData>')
    for row_index in range(1, max_row + 1):
        row_cells = []
        for col_index in range(1, max_col + 1):
            value = cell_map.get((row_index, col_index))
            if value is None or value == "":
                continue
            ref = f'{column_name(col_index)}{row_index}'
            style_name = style_map.get((row_index, col_index), "default")
            style_id = STYLE_NAME_TO_ID.get(style_name, 0)
            style_attr = f' s="{style_id}"' if style_id else ""
            if isinstance(value, Decimal):
                row_cells.append(f'<c r="{ref}"{style_attr}><v>{format(value, "f")}</v></c>')
            elif isinstance(value, (int, float)) and not isinstance(value, bool):
                row_cells.append(f'<c r="{ref}"{style_attr}><v>{value}</v></c>')
            else:
                row_cells.append(f'<c r="{ref}" t="inlineStr"{style_attr}><is><t>{escape_xml(value)}</t></is></c>')
        if row_cells or row_index in row_heights:
            height_attr = f' ht="{row_heights[row_index]}" customHeight="1"' if row_index in row_heights else ""
            lines.append(f'<row r="{row_index}"{height_attr}>' + ''.join(row_cells) + '</row>')
    lines.append('</sheetData>')
    if include_drawing:
        lines.append('<drawing r:id="rId1"/>')
    lines.append('</worksheet>')
    return ''.join(lines).encode('utf-8')


STYLES_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="8">
    <font><sz val="11"/><color rgb="FF000000"/><name val="Calibri"/></font>
    <font><b/><sz val="11"/><color rgb="FF000000"/><name val="Calibri"/></font>
    <font><b/><sz val="11"/><color rgb="FF006100"/><name val="Calibri"/></font>
    <font><sz val="11"/><color rgb="FF000000"/><name val="Calibri"/></font>
    <font><sz val="11"/><color rgb="FFFF0000"/><name val="Calibri"/></font>
    <font><sz val="11"/><color rgb="FF000099"/><name val="Calibri"/></font>
    <font><b/><sz val="11"/><color rgb="FF000000"/><name val="Calibri"/></font>
    <font><sz val="10"/><color rgb="FF000000"/><name val="Calibri"/></font>
  </fonts>
  <fills count="4">
    <fill><patternFill patternType="none"/></fill>
    <fill><patternFill patternType="gray125"/></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFC6EFCE"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFFFFFFF"/><bgColor indexed="64"/></patternFill></fill>
  </fills>
  <borders count="2">
    <border><left/><right/><top/><bottom/><diagonal/></border>
    <border><left style="thin"><color auto="1"/></left><right style="thin"><color auto="1"/></right><top style="thin"><color auto="1"/></top><bottom style="thin"><color auto="1"/></bottom><diagonal/></border>
  </borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="10">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
    <xf numFmtId="0" fontId="1" fillId="0" borderId="0" xfId="0" applyFont="1"/>
    <xf numFmtId="0" fontId="2" fillId="2" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1"/>
    <xf numFmtId="0" fontId="3" fillId="2" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1"/>
    <xf numFmtId="0" fontId="4" fillId="3" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1"/>
    <xf numFmtId="0" fontId="5" fillId="3" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1"/>
    <xf numFmtId="0" fontId="6" fillId="2" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1"/>
    <xf numFmtId="0" fontId="7" fillId="0" borderId="0" xfId="0" applyFont="1"/>
    <xf numFmtId="0" fontId="5" fillId="3" borderId="0" xfId="0" applyFont="1" applyFill="1"/>
    <xf numFmtId="0" fontId="4" fillId="3" borderId="0" xfId="0" applyFont="1" applyFill="1"/>
  </cellXfs>
  <cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
</styleSheet>"""


def build_drawing_xml(image):
    width = int(image.get("width", 220) * 9525)
    height = int(image.get("height", 70) * 9525)
    from_col = int(image.get("from_col", 0))
    from_row = int(image.get("from_row", 0))
    return f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <xdr:oneCellAnchor editAs="oneCell">
    <xdr:from>
      <xdr:col>{from_col}</xdr:col>
      <xdr:colOff>0</xdr:colOff>
      <xdr:row>{from_row}</xdr:row>
      <xdr:rowOff>0</xdr:rowOff>
    </xdr:from>
    <xdr:ext cx="{width}" cy="{height}"/>
    <xdr:pic>
      <xdr:nvPicPr>
        <xdr:cNvPr id="1" name="Picture 1" descr="Company logo"/>
        <xdr:cNvPicPr>
          <a:picLocks noChangeAspect="1"/>
        </xdr:cNvPicPr>
      </xdr:nvPicPr>
      <xdr:blipFill>
        <a:blip r:embed="rId1" cstate="print"/>
        <a:stretch><a:fillRect/></a:stretch>
      </xdr:blipFill>
      <xdr:spPr>
        <a:xfrm>
          <a:off x="0" y="0"/>
          <a:ext cx="{width}" cy="{height}"/>
        </a:xfrm>
        <a:prstGeom prst="rect"><a:avLst/></a:prstGeom>
      </xdr:spPr>
    </xdr:pic>
    <xdr:clientData/>
  </xdr:oneCellAnchor>
</xdr:wsDr>'''.encode('utf-8')


def build_xlsx(cell_map, sheet_name="Blad1", style_map=None, column_widths=None, row_heights=None, image=None):
    output = io.BytesIO()
    include_image = bool(image and image.get("bytes"))
    image_ext = (image or {}).get("ext", "png")
    with ZipFile(output, "w", ZIP_DEFLATED) as zf:
        content_types = ["""<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>
<Types xmlns=\"http://schemas.openxmlformats.org/package/2006/content-types\">
  <Default Extension=\"rels\" ContentType=\"application/vnd.openxmlformats-package.relationships+xml\"/>
  <Default Extension=\"xml\" ContentType=\"application/xml\"/>
"""]
        if include_image:
            content_types.append(f'  <Default Extension="{image_ext}" ContentType="image/{"jpeg" if image_ext in ("jpg", "jpeg") else image_ext}"/>\n')
        content_types.append("""  <Override PartName=\"/xl/workbook.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml\"/>
  <Override PartName=\"/xl/worksheets/sheet1.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml\"/>
  <Override PartName=\"/xl/styles.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml\"/>
""")
        if include_image:
            content_types.append('  <Override PartName="/xl/drawings/drawing1.xml" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>\n')
        content_types.append("""  <Override PartName=\"/docProps/core.xml\" ContentType=\"application/vnd.openxmlformats-package.core-properties+xml\"/>
  <Override PartName=\"/docProps/app.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.extended-properties+xml\"/>
</Types>""")
        zf.writestr('[Content_Types].xml', ''.join(content_types))
        zf.writestr("_rels/.rels", f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="{NS_PKG_REL}">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>
</Relationships>''')
        zf.writestr("docProps/core.xml", '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:dcmitype="http://purl.org/dc/dcmitype/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <dc:creator>SnappySjaak UKdocs</dc:creator>
  <cp:lastModifiedBy>SnappySjaak UKdocs</cp:lastModifiedBy>
  <dcterms:created xsi:type="dcterms:W3CDTF">2026-06-18T00:00:00Z</dcterms:created>
  <dcterms:modified xsi:type="dcterms:W3CDTF">2026-06-18T00:00:00Z</dcterms:modified>
</cp:coreProperties>''')
        zf.writestr("docProps/app.xml", f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
  <Application>SnappySjaak UKdocs</Application>
  <TitlesOfParts><vt:vector size="1" baseType="lpstr"><vt:lpstr>{escape_xml(sheet_name)}</vt:lpstr></vt:vector></TitlesOfParts>
</Properties>''')
        zf.writestr("xl/workbook.xml", f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="{NS_MAIN}" xmlns:r="{NS_REL}">
  <sheets><sheet name="{escape_xml(sheet_name)}" sheetId="1" r:id="rId1"/></sheets>
</workbook>''')
        zf.writestr("xl/_rels/workbook.xml.rels", f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="{NS_PKG_REL}">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>''')
        zf.writestr("xl/styles.xml", STYLES_XML)
        zf.writestr("xl/worksheets/sheet1.xml", build_sheet_xml(cell_map, style_map=style_map, column_widths=column_widths, row_heights=row_heights, include_drawing=include_image))
        if include_image:
            zf.writestr("xl/worksheets/_rels/sheet1.xml.rels", f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="{NS_PKG_REL}">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing1.xml"/>
</Relationships>''')
            zf.writestr("xl/drawings/drawing1.xml", build_drawing_xml(image))
            zf.writestr("xl/drawings/_rels/drawing1.xml.rels", f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="{NS_PKG_REL}">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="../media/image1.{image_ext}"/>
</Relationships>''')
            zf.writestr(f"xl/media/image1.{image_ext}", image["bytes"])
    return output.getvalue()


PUBLIC_DIR = Path(__file__).resolve().parent.parent / "public"


def resolve_template_path(name):
    template_name = clean_text(name)
    if not template_name:
        return None
    candidate = PUBLIC_DIR / template_name.lstrip("/")
    if candidate.exists() and candidate.is_file():
        return candidate
    return None


def load_template_sheet(template_name, kind):
    if not clean_text(template_name):
        return None, None
    if load_workbook is None:
        raise RuntimeError(f"openpyxl is required to generate {kind} files from templates")
    template_path = resolve_template_path(template_name)
    if not template_path:
        raise RuntimeError(f'{kind.title()} template not found: {template_name}')
    workbook = load_workbook(template_path)
    return workbook, workbook[workbook.sheetnames[0]]


def workbook_to_bytes(workbook):
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def excel_value(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, str):
        return strip_invalid_xml_chars(value)
    return value


def set_sheet_value(sheet, row, col, value):
    sheet.cell(row=row, column=col).value = excel_value(value)


def clear_sheet_range(sheet, start_row, end_row, start_col, end_col):
    if end_row < start_row or end_col < start_col:
        return
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            sheet.cell(row=row, column=col).value = None


def copy_row_format(sheet, source_row, target_row, max_col):
    if source_row <= 0 or target_row <= 0 or source_row == target_row:
        return
    source_dimensions = sheet.row_dimensions[source_row]
    target_dimensions = sheet.row_dimensions[target_row]
    target_dimensions.height = source_dimensions.height
    target_dimensions.hidden = source_dimensions.hidden
    for col in range(1, max_col + 1):
        source_cell = sheet.cell(row=source_row, column=col)
        target_cell = sheet.cell(row=target_row, column=col)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)


def copy_block_format(sheet, source_row, target_row, row_count, max_col):
    for offset in range(row_count):
        copy_row_format(sheet, source_row + offset, target_row + offset, max_col)


def row_contains_terms(sheet, row_index, terms):
    values = [clean_text(sheet.cell(row=row_index, column=col).value).lower() for col in range(1, sheet.max_column + 1)]
    return all(any(term in value for value in values) for term in terms)


def find_row_with_terms(sheet, terms, start_row=1):
    normalized_terms = [clean_text(term).lower() for term in terms if clean_text(term)]
    for row_index in range(max(1, start_row), sheet.max_row + 1):
        if row_contains_terms(sheet, row_index, normalized_terms):
            return row_index
    return 0


def find_cell_containing(sheet, needle, start_row=1):
    wanted = clean_text(needle).lower()
    for row_index in range(max(1, start_row), sheet.max_row + 1):
        for col_index in range(1, sheet.max_column + 1):
            value = clean_text(sheet.cell(row=row_index, column=col_index).value).lower()
            if wanted and wanted in value:
                return row_index, col_index
    return 0, 0


def format_shipment_date(date_text):
    text = clean_text(date_text)
    if not text:
        return ""
    try:
        return datetime.strptime(text, "%Y-%m-%d").strftime("%d-%m-%Y")
    except ValueError:
        return text


def write_invoice_template(workbook, sheet, analysis, category_code):
    category = next(item for item in analysis["categories"] if item["code"] == category_code)
    customer = analysis["customer"]
    shipment = analysis["shipment"]
    company = analysis["company"]

    customer_lines = [customer.get("customer_name", "")]
    customer_lines.extend(str(customer.get("customer_address", "") or "").splitlines())
    customer_lines.append(f'VAT NR {customer.get("vat_number", "")}')
    customer_lines.append(f'EORI NR {customer.get("importer_number") or customer.get("eori_number") or ""}')
    customer_lines.append(customer.get("importer_number") or "")
    while len(customer_lines) < 6:
        customer_lines.append("")
    customer_start_row = 10
    clear_sheet_range(sheet, customer_start_row, customer_start_row + 5, 2, 3)
    for offset, value in enumerate(customer_lines[:6], start=customer_start_row):
        set_sheet_value(sheet, offset, 2, value)

    date_row, date_col = find_cell_containing(sheet, "Date")
    invoice_row, invoice_col = find_cell_containing(sheet, "Invoice nr")
    licence_row, licence_col = find_cell_containing(sheet, "Licence Truck")
    if not licence_row:
        licence_row, licence_col = find_cell_containing(sheet, "License Truck")
    delivery_row, delivery_col = find_cell_containing(sheet, "Delivery T")

    date_col = date_col or 1
    invoice_col = invoice_col or date_col
    licence_col = licence_col or date_col
    delivery_col = delivery_col or date_col

    info_start_row = 17
    clear_sheet_range(sheet, 16, 21, 1, 6)
    set_sheet_value(sheet, info_start_row, date_col, "Date :")
    set_sheet_value(sheet, info_start_row, date_col + 1, format_shipment_date(shipment.get("shipment_date_excel")))
    set_sheet_value(sheet, info_start_row + 1, invoice_col, "Invoice nr :")
    set_sheet_value(sheet, info_start_row + 1, invoice_col + 1, category.get("invoice_number", ""))
    set_sheet_value(sheet, info_start_row + 1, invoice_col + 2, "custom summary")
    set_sheet_value(sheet, info_start_row + 2, licence_col, "Licence Truck :")
    set_sheet_value(sheet, info_start_row + 2, licence_col + 1, shipment.get("trailer_number", ""))
    set_sheet_value(sheet, info_start_row + 3, delivery_col, "Delivery Terms :")
    set_sheet_value(sheet, info_start_row + 3, delivery_col + 1, shipment.get("delivery_terms", ""))

    table_header_row = 22
    hs_header_row = find_row_with_terms(sheet, ["goods description", "packages"], table_header_row + 1) or (table_header_row + 4)
    original_footer_row = find_row_with_terms(sheet, ["vat nr"], hs_header_row + 1) or max(sheet.max_row - 7, hs_header_row + 8)
    footer_row = original_footer_row

    invoice_start_row = table_header_row + 1
    existing_invoice_space = max(hs_header_row - invoice_start_row - 1, 0)
    needed_invoice_space = len(category["invoice_rows"]) + 1
    if needed_invoice_space > existing_invoice_space:
        extra_rows = needed_invoice_space - existing_invoice_space
        template_row = invoice_start_row
        sheet.insert_rows(hs_header_row, extra_rows)
        for index in range(extra_rows):
            copy_row_format(sheet, template_row, hs_header_row + index, 10)
        hs_header_row += extra_rows
        footer_row += extra_rows

    clear_sheet_range(sheet, invoice_start_row, hs_header_row - 1, 1, 10)
    headers = ["classificationType TARIC", "Goods description", "Origin", "Quantity", "Gross kg", "Net kg", "Packages", "Value"]
    for offset, value in enumerate(headers, start=2):
        set_sheet_value(sheet, table_header_row, offset, value)

    row_number = invoice_start_row
    for line in category["invoice_rows"]:
        set_sheet_value(sheet, row_number, 2, line["commodity_code"])
        set_sheet_value(sheet, row_number, 3, line["description"])
        set_sheet_value(sheet, row_number, 4, line["origin"])
        set_sheet_value(sheet, row_number, 5, line["quantity"])
        set_sheet_value(sheet, row_number, 6, line["gross_kg"])
        set_sheet_value(sheet, row_number, 7, line["net_kg"])
        set_sheet_value(sheet, row_number, 8, line["packages"])
        set_sheet_value(sheet, row_number, 9, line["customs_value"])
        row_number += 1

    hs_header_row = max(hs_header_row, row_number + 1)
    clear_sheet_range(sheet, row_number, hs_header_row - 1, 1, 10)

    hs_headers = ["classificationTyp HS", "Goods description", "Quantity", "gros kg", "net kg", "Packages", "Value"]
    for offset, value in enumerate(hs_headers, start=3):
        set_sheet_value(sheet, hs_header_row, offset, value)

    hs_start_row = hs_header_row + 1
    data_style_row = hs_start_row
    existing_hs_space = max(footer_row - hs_start_row - 1, 0)
    needed_hs_space = len(category["hs_summary_rows"]) + 2
    if needed_hs_space > existing_hs_space:
        extra_rows = needed_hs_space - existing_hs_space
        template_row = hs_start_row
        sheet.insert_rows(footer_row, extra_rows)
        for index in range(extra_rows):
            copy_row_format(sheet, template_row, footer_row + index, 10)
        footer_row += extra_rows

    clear_sheet_range(sheet, hs_start_row, footer_row - 1, 3, 10)
    row_number = hs_start_row
    for line in category["hs_summary_rows"]:
        copy_row_format(sheet, data_style_row, row_number, 10)
        set_sheet_value(sheet, row_number, 3, normalize_invoice_hs_code(line["hs_code"]))
        set_sheet_value(sheet, row_number, 4, line["description"])
        set_sheet_value(sheet, row_number, 5, line["quantity"])
        set_sheet_value(sheet, row_number, 6, line["gross_kg"])
        set_sheet_value(sheet, row_number, 7, line["net_kg"])
        set_sheet_value(sheet, row_number, 8, line["packages"])
        set_sheet_value(sheet, row_number, 9, line["customs_value"])
        row_number += 1

    totals_row = row_number
    copy_row_format(sheet, data_style_row, totals_row, 10)
    set_sheet_value(sheet, totals_row, 6, "TOTALS")
    set_sheet_value(sheet, totals_row, 7, category["totals"]["net_kg"])
    set_sheet_value(sheet, totals_row, 9, category["totals"]["customs_value"])

    footer_row = max(totals_row + 4, hs_header_row + 6)
    if original_footer_row > 0 and footer_row != original_footer_row:
        copy_block_format(sheet, original_footer_row, footer_row, 8, 10)
        clear_sheet_range(sheet, original_footer_row, original_footer_row + 7, 3, 8)
    clear_sheet_range(sheet, footer_row, footer_row + 7, 3, 8)
    set_sheet_value(sheet, footer_row, 3, company.get("company_name", ""))
    set_sheet_value(sheet, footer_row, 7, f'VAT nr : {company.get("vat_number", "")}')
    address_lines = str(company.get("address", "") or "").splitlines()
    if len(address_lines) > 0:
        set_sheet_value(sheet, footer_row + 1, 3, address_lines[0])
    if len(address_lines) > 1:
        set_sheet_value(sheet, footer_row + 2, 3, address_lines[1])
    set_sheet_value(sheet, footer_row + 1, 7, f'EORI nr:{company.get("eori_number", "")}')
    set_sheet_value(sheet, footer_row + 2, 7, f'Chamber of Commerce : {company.get("chamber_of_commerce_number", "")}')
    set_sheet_value(sheet, footer_row + 3, 3, f'tel {company.get("phone", "")}'.strip())
    set_sheet_value(sheet, footer_row + 3, 7, f'IBAN : {company.get("iban", "")}')
    set_sheet_value(sheet, footer_row + 4, 3, f'email : {company.get("email", "")}'.strip())
    set_sheet_value(sheet, footer_row + 4, 7, f'BIC/SWIFT : {company.get("bic_swift", "")}')
    set_sheet_value(sheet, footer_row + 5, 3, f'web :  {company.get("website", "")}'.strip())
    set_sheet_value(sheet, footer_row + 5, 7, f'rex registration : {company.get("rex_registration", "")}')
    set_sheet_value(sheet, footer_row + 6, 3, company.get("default_footer_text", ""))
    set_sheet_value(sheet, footer_row + 7, 3, company.get("preferential_origin_declaration", ""))

    return workbook_to_bytes(workbook)


def write_export_template(workbook, sheet, analysis):
    shipment = analysis["shipment"]
    customer = analysis.get("customer") or {}

    header_row = find_row_with_terms(sheet, ["reference", "owner", "regulation"], 1) or 1
    value_row = header_row + 1
    goods_header_row = find_row_with_terms(sheet, ["goods description", "commodity code", "net weight"], value_row) or 3
    data_start_row = goods_header_row + 1

    header_values = [
        "Reference", "Owner", "Regulation", "Country of destination", "Total gross mass", "Total number of packages", "Location", "Marks and numbers", "Container number", "Border transport mode", "Border transport nationality", "Delivery terms", "Delivery terms city", "Customs office of exit"
    ]
    for index, value in enumerate(header_values, start=1):
        set_sheet_value(sheet, header_row, index, value)

    value_values = [
        shipment.get("export_header_reference") or shipment["reference_line"], shipment["owner"], shipment["regulation"], shipment["destination_country"], f"=SUM(G{data_start_row}:G99)", f"=SUM(F{data_start_row}:F99)", shipment["location"], shipment["marks_and_numbers"], shipment["container_number"], shipment["border_transport_mode"], shipment["border_transport_nationality"], shipment["delivery_terms"], shipment["delivery_terms_city"], shipment["customs_office_of_exit"],
    ]
    for index, value in enumerate(value_values, start=1):
        set_sheet_value(sheet, value_row, index, value)

    goods_headers = ["Goods description", "Commodity code", "Net weight", "Quantity", "Customs value", "Ctns per regel", "Bruto per regel", "oorsprong", "Certificate Origin", "Bio Certificate", "KCB Number", "Phyto number"]
    for index, value in enumerate(goods_headers, start=1):
        set_sheet_value(sheet, goods_header_row, index, value)

    existing_max_row = max(sheet.max_row, data_start_row + len(analysis["export_rows"]) + 5)
    clear_sheet_range(sheet, data_start_row, existing_max_row, 1, 12)
    template_row = data_start_row
    for index, row in enumerate(analysis["export_rows"]):
        row_number = data_start_row + index
        if row_number > sheet.max_row:
            copy_row_format(sheet, template_row, row_number, 15)
        set_sheet_value(sheet, row_number, 1, row["description"])
        set_sheet_value(sheet, row_number, 2, row["commodity_code"])
        set_sheet_value(sheet, row_number, 3, row["net_kg"])
        set_sheet_value(sheet, row_number, 4, row["quantity"])
        set_sheet_value(sheet, row_number, 5, row["customs_value"])
        set_sheet_value(sheet, row_number, 6, row["packages"])
        set_sheet_value(sheet, row_number, 7, row["gross_kg"])
        set_sheet_value(sheet, row_number, 8, row["origin"])

    detail_values = {
        4: {13: "Additional Information"},
        5: {13: "Importer", 14: shipment["customer_importer_number"]},
        6: {14: customer.get("customer_name", "")},
        7: {13: "Invoice number", 14: shipment["invoice_numbers"]},
        8: {13: "Trailer number", 14: shipment["trailer_number"]},
        9: {13: "Vessel", 14: shipment["vessel"]},
        11: {13: "Port of UK Arrival", 14: shipment["uk_arrival_port"]},
        12: {13: "Currency of invoice", 14: shipment["currency"]},
        13: {15: "Currency"},
        14: {13: "Freight costs", 14: shipment["freight_costs"], 15: shipment["currency"]},
        15: {13: "Insurance", 14: shipment["insurance"]},
        16: {13: "Inland freight", 15: shipment["currency"]},
    }
    for row_index, columns in detail_values.items():
        for col_index, value in columns.items():
            set_sheet_value(sheet, row_index, col_index, value)

    return workbook_to_bytes(workbook)


def build_export_workbook_raw(analysis):
    shipment = analysis["shipment"]
    customer = analysis.get("customer") or {}
    company = analysis.get("company") or {}
    cells = rows_to_dense_map([
        ["Reference", "Owner", "Regulation", "Country of destination", "Total gross mass", "Total number of packages", "Location", "Marks and numbers", "Container number", "Border transport mode", "Border transport nationality", "Delivery terms", "Delivery terms city", "Customs office of exit"],
        [shipment.get("export_header_reference") or shipment["reference_line"], shipment["owner"], shipment["regulation"], shipment["destination_country"], analysis["combined_totals"]["gross_kg"], analysis["combined_totals"]["packages"], shipment["location"], shipment["marks_and_numbers"], shipment["container_number"], shipment["border_transport_mode"], shipment["border_transport_nationality"], shipment["delivery_terms"], shipment["delivery_terms_city"], shipment["customs_office_of_exit"]],
        ["Goods description", "Commodity code", "Net weight", "Quantity", "Customs value", "Ctns per regel", "Bruto per regel", "oorsprong", "Certificate Origin", "Bio Certificate", "KCB Number", "Phyto number"],
    ])
    style_map = {}
    column_widths = {
        1: 22.73,
        2: 15.27,
        3: 10.86,
        4: 18.86,
        5: 18.40,
        6: 21.13,
        7: 15.13,
        8: 15.60,
        9: 15.60,
        10: 19.13,
        11: 23.40,
        12: 14.73,
        13: 18.86,
        14: 21.86,
        15: 19.13,
    }
    row_heights = {1: 18, 2: 18, 3: 18}

    for col in range(1, 15):
        style_map[(1, col)] = "export_header"
    for col in range(1, 13):
        style_map[(3, col)] = "table_header"
    cells[(1, 15)] = " "

    export_row_start = 4
    for index, row in enumerate(analysis["export_rows"]):
        row_number = export_row_start + index
        cells[(row_number, 1)] = row["description"]
        cells[(row_number, 2)] = row["commodity_code"]
        cells[(row_number, 3)] = row["net_kg"]
        cells[(row_number, 4)] = row["quantity"]
        cells[(row_number, 5)] = row["customs_value"]
        cells[(row_number, 6)] = row["packages"]
        cells[(row_number, 7)] = row["gross_kg"]
        cells[(row_number, 8)] = row["origin"]
        for col in range(1, 13):
            if col <= 8:
                style_map[(row_number, col)] = "export_goods"
            else:
                cells[(row_number, col)] = " "
                style_map[(row_number, col)] = "default"

    details = [
        (4, 13, "Additional Information"),
        (5, 13, "Importer"), (5, 14, shipment["customer_importer_number"]),
        (6, 14, customer.get("customer_name", "")),
        (7, 13, "Invoice number"), (7, 14, shipment["invoice_numbers"]),
        (8, 13, "Trailer number"), (8, 14, shipment["trailer_number"]),
        (9, 13, "Vessel"), (9, 14, shipment["vessel"]),
        (11, 13, "Port of UK Arrival"), (11, 14, shipment["uk_arrival_port"]),
        (12, 13, "Currency of invoice"), (12, 14, shipment["currency"]),
        (13, 15, "Currency"),
        (14, 13, "Freight costs"), (14, 14, shipment["freight_costs"]), (14, 15, shipment["currency"]),
        (15, 13, "Insurance"), (15, 14, shipment["insurance"]),
        (16, 13, "Inland freight"), (16, 15, shipment["currency"]),
    ]
    for row, col, value in details:
        if value:
            cells[(row, col)] = value
            if col == 13 or (row == 13 and col == 15):
                style_map[(row, col)] = "detail_label"
            elif col == 14:
                style_map[(row, col)] = "detail_value"
            else:
                style_map[(row, col)] = "detail_value_plain"

    for row in range(4, max(17, export_row_start + len(analysis["export_rows"]))):
        for col in range(13, 16):
            if (row, col) not in cells:
                cells[(row, col)] = " "
                style_map[(row, col)] = "detail_label_plain" if col == 13 else ("detail_value" if col == 14 else "detail_value_plain")

    return build_xlsx(cells, sheet_name="Export", style_map=style_map, column_widths=column_widths, row_heights=row_heights)


def build_invoice_workbook_raw(analysis, category_code):
    category = next(item for item in analysis["categories"] if item["code"] == category_code)
    customer = analysis["customer"]
    shipment = analysis["shipment"]
    company = analysis["company"]
    cells = {}
    style_map = {}
    column_widths = {
        1: 4,
        2: 16,
        3: 21.86,
        4: 21.86,
        5: 12,
        6: 14.27,
        7: 14.27,
        8: 12,
        9: 14.27,
    }
    row_heights = {
        10: 12.95,
        11: 12.95,
        12: 12.95,
        13: 12.95,
        14: 12.95,
        15: 12.95,
        17: 12.95,
        18: 13.7,
        19: 13.7,
        20: 12.95,
        24: 12.95,
    }

    shipment_date = clean_text(shipment.get("shipment_date_excel"))
    if shipment_date:
        try:
            shipment_date = datetime.strptime(shipment_date, "%Y-%m-%d").strftime("%d-%m-%Y")
        except ValueError:
            pass

    customer_lines = [customer.get("customer_name", "")]
    customer_lines.extend(str(customer.get("customer_address", "") or "").splitlines())
    customer_lines.append(f'VAT NR {customer.get("vat_number", "")}')
    customer_lines.append(f'EORI NR {customer.get("importer_number") or customer.get("eori_number") or ""}')
    customer_lines.append(customer.get("importer_number") or "")
    for offset, value in enumerate(customer_lines[:6], start=10):
        cells[(offset, 2)] = value
        style_map[(offset, 2)] = "footer"

    cells[(17, 2)] = "Date :"
    cells[(17, 3)] = shipment_date
    cells[(18, 2)] = "Invoice nr :"
    cells[(18, 3)] = category["invoice_number"]
    cells[(18, 4)] = "custom summary"
    cells[(19, 2)] = "Licence Truck :"
    cells[(19, 3)] = shipment["trailer_number"]
    cells[(20, 2)] = "Delivery Terms :"
    cells[(20, 3)] = shipment["delivery_terms"]
    for row, col in [(17, 2), (18, 2), (18, 4), (19, 2), (20, 2)]:
        style_map[(row, col)] = "bold"
    for row, col in [(17, 3), (18, 3), (19, 3), (20, 3)]:
        style_map[(row, col)] = "footer"

    table_header_row = 24
    cells[(table_header_row, 1)] = " "
    style_map[(table_header_row, 1)] = "table_header"
    header_values = ["classificationType TARIC", "Goods description", "Origin", "Quantity", "Gross kg", "Net kg", "Packages", "Value"]
    for offset, value in enumerate(header_values, start=2):
        cells[(table_header_row, offset)] = value
        style_map[(table_header_row, offset)] = "table_header"
    row_number = table_header_row + 1
    for line in category["invoice_rows"]:
        cells[(row_number, 1)] = " "
        style_map[(row_number, 1)] = "footer"
        cells[(row_number, 2)] = line["commodity_code"]
        cells[(row_number, 3)] = line["description"]
        cells[(row_number, 4)] = line["origin"]
        cells[(row_number, 5)] = line["quantity"]
        cells[(row_number, 6)] = line["gross_kg"]
        cells[(row_number, 7)] = line["net_kg"]
        cells[(row_number, 8)] = line["packages"]
        cells[(row_number, 9)] = line["customs_value"]
        for col in range(2, 10):
            style_map[(row_number, col)] = "footer"
        row_heights[row_number] = 10.9
        row_number += 1

    separator_row = row_number
    for col in [6, 7, 9]:
        cells[(separator_row, col)] = " "
        style_map[(separator_row, col)] = "footer"
    row_heights[separator_row] = 10.9

    hs_header_row = separator_row + 1
    hs_header_values = ["classificationTyp HS", "Goods description", "Quantity", "gros kg", "net kg", "Packages", "Value"]
    for offset, value in enumerate(hs_header_values, start=3):
        cells[(hs_header_row, offset)] = value
        style_map[(hs_header_row, offset)] = "table_header"
    row_heights[hs_header_row] = 10.9

    row_number = hs_header_row + 1
    for line in category["hs_summary_rows"]:
        cells[(row_number, 3)] = normalize_invoice_hs_code(line["hs_code"])
        cells[(row_number, 4)] = line["description"]
        cells[(row_number, 5)] = line["quantity"]
        cells[(row_number, 6)] = line["gross_kg"]
        cells[(row_number, 7)] = line["net_kg"]
        cells[(row_number, 8)] = line["packages"]
        cells[(row_number, 9)] = line["customs_value"]
        for col in range(3, 10):
            style_map[(row_number, col)] = "footer"
        row_heights[row_number] = 10.9
        row_number += 1

    totals_row = row_number + 1
    cells[(totals_row, 6)] = "TOTALS"
    cells[(totals_row, 7)] = category["totals"]["net_kg"]
    cells[(totals_row, 9)] = category["totals"]["customs_value"]
    style_map[(totals_row, 6)] = "bold"
    style_map[(totals_row, 7)] = "footer"
    style_map[(totals_row, 9)] = "footer"
    row_heights[totals_row] = 10.9

    footer_row = max(totals_row + 6, 36)
    cells[(footer_row, 3)] = company.get("company_name", "")
    style_map[(footer_row, 3)] = "bold"
    cells[(footer_row, 7)] = f'VAT nr : {company.get("vat_number", "")}'
    style_map[(footer_row, 7)] = "footer"
    address_lines = str(company.get("address", "") or "").splitlines()
    if len(address_lines) > 0:
        cells[(footer_row + 1, 3)] = address_lines[0]
    if len(address_lines) > 1:
        cells[(footer_row + 2, 3)] = address_lines[1]
    cells[(footer_row + 1, 7)] = f'EORI nr:{company.get("eori_number", "")}'
    cells[(footer_row + 2, 7)] = f'Chamber of Commerce : {company.get("chamber_of_commerce_number", "")}'
    cells[(footer_row + 3, 3)] = f'tel {company.get("phone", "")}'.strip()
    cells[(footer_row + 3, 7)] = f'IBAN : {company.get("iban", "")}'
    cells[(footer_row + 4, 3)] = f'email : {company.get("email", "")}'.strip()
    cells[(footer_row + 4, 7)] = f'BIC/SWIFT : {company.get("bic_swift", "")}'
    cells[(footer_row + 5, 3)] = f'web :  {company.get("website", "")}'.strip()
    cells[(footer_row + 5, 7)] = f'rex registration : {company.get("rex_registration", "")}'
    cells[(footer_row + 6, 3)] = company.get("default_footer_text", "")
    cells[(footer_row + 7, 3)] = company.get("preferential_origin_declaration", "")
    for row in range(footer_row + 1, footer_row + 8):
        row_heights[row] = 10.9
        if cells.get((row, 3)):
            style_map[(row, 3)] = "footer"
        if cells.get((row, 7)):
            style_map[(row, 7)] = "footer"

    logo_image = resolve_logo_image(company)
    if logo_image:
        logo_image = {**logo_image, "from_col": 1, "from_row": 1}
    return build_xlsx(cells, sheet_name="Invoice", style_map=style_map, column_widths=column_widths, row_heights=row_heights, image=logo_image)


def build_export_workbook(analysis):
    template_name = clean_text((analysis.get("templates") or {}).get("export_template_name"))
    if template_name:
        workbook, sheet = load_template_sheet(template_name, "export")
        return write_export_template(workbook, sheet, analysis)
    return build_export_workbook_raw(analysis)


def build_invoice_workbook(analysis, category_code):
    template_name = clean_text((analysis.get("templates") or {}).get("invoice_template_name"))
    if template_name:
        workbook, sheet = load_template_sheet(template_name, "invoice")
        return write_invoice_template(workbook, sheet, analysis, category_code)
    return build_invoice_workbook_raw(analysis, category_code)


def build_audit_workbook(analysis):
    rows = [
        ["Shipment reference", analysis["shipment"]["reference_line"]],
        ["Created at", datetime.utcnow().isoformat()],
        ["Final status", analysis["audit"]["final_status"]],
        [],
        ["Category", "Invoice", "Quantity", "Gross kg", "Net kg", "Packages", "Customs value", "Warnings"],
    ]
    for category in analysis["categories"]:
        rows.append([f'{category["code"]} {category["label"]}', category["invoice_number"], category["totals"]["quantity"], category["totals"]["gross_kg"], category["totals"]["net_kg"], category["totals"]["packages"], category["totals"]["customs_value"], "; ".join(category["warnings"])])
    rows.extend([[], ["Control summary"], ["Scope", "Group", "Field", "Dump value", "Invoice value", "Export value", "Invoice diff", "Export diff", "Status"]])
    for row in analysis["audit"].get("summary_rows", []):
        rows.append([row["scope"], row.get("group_label", ""), row["field"], row.get("dump_value", ""), row.get("invoice_value", ""), row.get("export_value", ""), row.get("invoice_difference", ""), row.get("export_difference", ""), row.get("status", "")])
    rows.extend([[], ["Raw checks"], ["Scope", "Category", "Field", "Left", "Right", "Difference", "Status"]])
    for check in analysis["audit"]["checks"]:
        rows.append([check["scope"], check.get("category", ""), check["field"], check["left"], check["right"], check["difference"], "PASS" if check["ok"] else "FAIL"])
    if analysis["audit"]["warnings"]:
        rows.extend([[], ["Warnings"]])
        for warning in analysis["audit"]["warnings"]:
            rows.append([warning["message"]])
    return build_xlsx(rows_to_dense_map(rows))


def analyze_payload(payload):
    invoice_numbers_by_category = parse_invoice_numbers(payload)
    uploaded_files = payload.get("uploaded_files") or {}
    records = []
    categories = []
    warnings = []
    blocking_warning_codes = set()
    invoice_number_seen = {}

    for category_code in CATEGORY_ORDER:
        file_entry = uploaded_files.get(category_code) or {}
        content_base64 = clean_text(file_entry.get("content_base64"))
        if not content_base64:
            continue
        raw_bytes = base64.b64decode(content_base64)
        rows = parse_xlsx_rows(raw_bytes)
        headers = rows[0] if rows else []
        alias_map = build_column_alias_map(payload, category_code)
        indexes = resolve_header_indexes(headers, alias_map)
        required = ["grossMassValue", "grossMassUnit", "netMassValue", "netMassUnit", "value", "originCountryCode", "classificationType", "classificationValue", "goodsDescriptionText", "quantityValue", "packages", "fullClassificationCode", "taricCode"]
        missing_required = [name for name in required if indexes.get(name) is None]
        category_warnings = []
        if missing_required:
            category_warnings.append(f'Missing required columns: {", ".join(missing_required)}')
            warnings.append({"code": "missing_required_columns", "category": category_code, "message": f'{category_code}: missing required columns: {", ".join(missing_required)}'})
            blocking_warning_codes.add("missing_required_columns")
            categories.append({"code": category_code, "label": CATEGORY_DEFINITIONS[category_code]["label"], "invoice_number": invoice_numbers_by_category.get(category_code, ""), "row_count": 0, "totals": {"quantity": ZERO, "gross_kg": ZERO, "net_kg": ZERO, "packages": ZERO, "customs_value": ZERO}, "warnings": category_warnings, "invoice_rows": [], "hs_summary_rows": []})
            continue
        invoice_number = invoice_numbers_by_category.get(category_code, "")
        if invoice_number:
            if invoice_number in invoice_number_seen and invoice_number_seen[invoice_number] != category_code:
                warnings.append({"code": "duplicate_invoice_number", "category": category_code, "message": f'Duplicate invoice number {invoice_number} used for {invoice_number_seen[invoice_number]} and {category_code}'})
                blocking_warning_codes.add("duplicate_invoice_number")
            invoice_number_seen[invoice_number] = category_code
        category_records = []
        for source_row in rows[1:]:
            description = clean_text(row_value(source_row, indexes, "classificationValue") or row_value(source_row, indexes, "goodsDescriptionText"))
            if not description:
                warnings.append({"code": "blank_classification_value", "category": category_code, "message": f'{category_code}: blank classification / description value'})
                blocking_warning_codes.add("blank_classification_value")
            commodity_code = clean_text(row_value(source_row, indexes, "fullClassificationCode") or row_value(source_row, indexes, "taricCode") or row_value(source_row, indexes, "classificationType"))
            if not commodity_code:
                warnings.append({"code": "missing_commodity_code", "category": category_code, "message": f'{category_code}: missing commodity code for description {description or "(blank)"}'})
                blocking_warning_codes.add("missing_commodity_code")
            origin = clean_text(row_value(source_row, indexes, "preferentialOriginCountryCode") or row_value(source_row, indexes, "originCountryCode"))
            if not origin:
                warnings.append({"code": "missing_origin", "category": category_code, "message": f'{category_code}: missing origin for description {description or "(blank)"}'})
                blocking_warning_codes.add("missing_origin")
            gross_kg, gross_warning = convert_weight_kg(row_value(source_row, indexes, "grossMassValue"), row_value(source_row, indexes, "grossMassUnit"))
            net_kg, net_warning = convert_weight_kg(row_value(source_row, indexes, "netMassValue"), row_value(source_row, indexes, "netMassUnit"))
            if gross_warning or net_warning:
                warnings.append({"code": "unsupported_weight_unit", "category": category_code, "message": f'{category_code}: {gross_warning or net_warning}'})
                blocking_warning_codes.add("unsupported_weight_unit")
            quantity = dec(row_value(source_row, indexes, "quantityValue"))
            packages = dec(row_value(source_row, indexes, "packages"))
            customs_value = dec(row_value(source_row, indexes, "value"))
            if quantity == ZERO and clean_text(row_value(source_row, indexes, "quantityValue")) == "":
                warnings.append({"code": "missing_quantity", "category": category_code, "message": f'{category_code}: missing quantity for {description or commodity_code}'})
                blocking_warning_codes.add("missing_quantity")
            if packages == ZERO and clean_text(row_value(source_row, indexes, "packages")) == "":
                warnings.append({"code": "missing_packages", "category": category_code, "message": f'{category_code}: missing packages for {description or commodity_code}'})
                blocking_warning_codes.add("missing_packages")
            if customs_value == ZERO and clean_text(row_value(source_row, indexes, "value")) == "":
                warnings.append({"code": "missing_value", "category": category_code, "message": f'{category_code}: missing value for {description or commodity_code}'})
                blocking_warning_codes.add("missing_value")
            if quantity < ZERO or packages < ZERO or customs_value < ZERO or gross_kg < ZERO or net_kg < ZERO:
                warnings.append({"code": "negative_value", "category": category_code, "message": f'{category_code}: negative values detected for {description or commodity_code}'})
                blocking_warning_codes.add("negative_value")
            record = {
                "category": category_code,
                "category_label": CATEGORY_DEFINITIONS[category_code]["label"],
                "invoice_number": invoice_number,
                "hs_code": clean_text(row_value(source_row, indexes, "classificationType")),
                "commodity_code": commodity_code,
                "description": description,
                "origin": origin,
                "quantity": quantity,
                "gross_kg": gross_kg,
                "net_kg": net_kg,
                "packages": packages,
                "customs_value": customs_value,
            }
            category_records.append(record)
            records.append(record)
        category_totals = sum_dicts(category_records, ["quantity", "gross_kg", "net_kg", "packages", "customs_value"])
        invoice_rows = aggregate_rows(category_records, ["hs_code", "commodity_code", "description", "origin"])
        hs_summary_rows = aggregate_rows(category_records, ["hs_code", "description"])
        categories.append({"code": category_code, "label": CATEGORY_DEFINITIONS[category_code]["label"], "invoice_number": invoice_number, "row_count": len(category_records), "totals": category_totals, "warnings": category_warnings, "invoice_rows": invoice_rows, "hs_summary_rows": hs_summary_rows})

    export_rows = aggregate_rows(records, ["description", "commodity_code", "origin"])
    control_pivot_rows = aggregate_rows(records, ["description"])
    combined_totals = sum_dicts(records, ["quantity", "gross_kg", "net_kg", "packages", "customs_value"])
    category_totals = {item["code"]: item["totals"] for item in categories}
    invoice_totals = {item["code"]: sum_dicts(item["invoice_rows"], ["quantity", "gross_kg", "net_kg", "packages", "customs_value"]) for item in categories}
    combined_export_totals = sum_dicts(export_rows, ["quantity", "gross_kg", "net_kg", "packages", "customs_value"])
    checks, audit_summary_rows = build_audit_checks(category_totals, invoice_totals, combined_totals, combined_export_totals, payload)
    check_failures = [check for check in checks if not check["ok"]]
    final_status = "PASS" if not blocking_warning_codes and not check_failures else "FAIL"
    combined_invoice_numbers = clean_text(payload.get("invoice_numbers")) or "/".join(filter(None, [item.get("invoice_number") for item in categories]))
    reference_line = clean_text(payload.get("export_reference")) or combined_invoice_numbers
    customer = next((item for item in (payload.get("customers") or []) if clean_text(item.get("id")) == clean_text(payload.get("customer_id"))), {})
    company = payload.get("company_settings") or {}
    shipment = {
        "reference_line": reference_line,
        "export_header_reference": clean_text(payload.get("invoice_numbers")) + "  " + clean_text(payload.get("shipment_date", "")).replace("-", "") + " " + clean_text(payload.get("trailer_number")) if clean_text(payload.get("invoice_numbers")) and clean_text(payload.get("shipment_date")) and clean_text(payload.get("trailer_number")) else reference_line,
        "owner": clean_text(payload.get("owner")) or clean_text(company.get("eori_number")) or clean_text(company.get("vat_number")) or clean_text(company.get("company_name")),
        "regulation": clean_text(payload.get("regulation")) or "Export",
        "destination_country": clean_text(payload.get("destination_country")) or "GB",
        "location": clean_text(payload.get("location")),
        "marks_and_numbers": clean_text(payload.get("marks_and_numbers")) or reference_line,
        "container_number": clean_text(payload.get("container_number")),
        "border_transport_mode": clean_text(payload.get("border_transport_mode")) or "Road",
        "border_transport_nationality": clean_text(payload.get("border_transport_nationality")) or "NL",
        "delivery_terms": clean_text(payload.get("delivery_terms")),
        "delivery_terms_city": clean_text(payload.get("delivery_terms_city")),
        "customs_office_of_exit": clean_text(payload.get("customs_office_of_exit")),
        "trailer_number": clean_text(payload.get("trailer_number")),
        "invoice_numbers": combined_invoice_numbers,
        "vessel": clean_text(payload.get("vessel")),
        "uk_arrival_port": clean_text(payload.get("uk_arrival_port")),
        "currency": clean_text(payload.get("currency")) or "GBP",
        "freight_costs": clean_text(payload.get("freight_costs")),
        "insurance": clean_text(payload.get("insurance")),
        "customer_importer_number": clean_text(payload.get("importer")) or clean_text(customer.get("importer_number") or customer.get("eori_number")),
        "shipment_date_excel": clean_text(payload.get("shipment_date")),
    }
    return {
        "categories": categories,
        "export_rows": export_rows,
        "control_pivot_rows": control_pivot_rows,
        "combined_totals": combined_totals,
        "audit": {"final_status": final_status, "checks": checks, "summary_rows": audit_summary_rows, "warnings": warnings, "blocking_warning_codes": sorted(blocking_warning_codes)},
        "shipment": shipment,
        "customer": customer,
        "company": company,
        "templates": payload.get("templates") or {},
    }


def generate_files(analysis):
    shipment = analysis["shipment"]
    date_bits = clean_text(shipment.get("shipment_date_excel")).replace("-", " ")
    trailer = clean_text(shipment.get("trailer_number")) or "truck"
    export_reference = clean_text(shipment.get("reference_line")) or "reference"
    files = [{"name": f'{date_bits} {trailer} {export_reference}.xlsx'.replace("  ", " ").strip(), "content_base64": base64.b64encode(build_export_workbook(analysis)).decode("ascii"), "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "kind": "export"}]
    customer_name = clean_text((analysis["customer"] or {}).get("customer_name")) or "customer"
    for category in analysis["categories"]:
        if not category["row_count"]:
            continue
        files.append({"name": f'invoice {category["invoice_number"] or category["code"]} {CATEGORY_DEFINITIONS[category["code"]]["slug"]} {customer_name}.xlsx', "content_base64": base64.b64encode(build_invoice_workbook(analysis, category["code"])).decode("ascii"), "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "kind": "invoice", "category": category["code"]})
    files.append({"name": f'audit {export_reference}.xlsx', "content_base64": base64.b64encode(build_audit_workbook(analysis)).decode("ascii"), "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "kind": "audit"})
    return files


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: ukdocs_worker.py <analyze|generate|import-examples>")
    command = sys.argv[1]
    payload = json.loads(sys.stdin.read() or "{}")
    if command == "import-examples":
        print(json.dumps(json_decimal(import_example_payload(payload))))
        return
    analysis = analyze_payload(payload)
    if command == "analyze":
        print(json.dumps(json_decimal(analysis)))
        return
    if command == "generate":
        if analysis["audit"]["final_status"] != "PASS":
            raise SystemExit("UKdocs audit failed. Fix validation warnings and mismatches before generating documents.")
        print(json.dumps({"analysis": json_decimal(analysis), "files": generate_files(analysis)}))
        return
    raise SystemExit(f"Unknown command: {command}")


if __name__ == "__main__":
    main()
