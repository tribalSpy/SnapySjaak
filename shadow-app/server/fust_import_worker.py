import argparse
import json
import re
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


CARRIER_GROUPS = [
    {"sheetLabel": "Breewel", "siteKlant": "Breewel"},
    {"sheetLabel": "ML Express", "siteKlant": "ML Express Parijs"},
    {"sheetLabel": "De Wit", "siteKlant": "De Wit"},
    {"sheetLabel": "De Wit 2", "siteKlant": "De Wit 2"},
]

CARRIER_NAME_MAP = {item["sheetLabel"].lower(): item["siteKlant"] for item in CARRIER_GROUPS}

DATE_RE = re.compile(r"(\d{2})-(\d{2})-(\d{4})")


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def to_number(value):
    text = clean_text(value).replace(",", ".")
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def load_rows(input_path: Path):
    data = input_path.read_bytes()
    workbook = load_workbook(filename=BytesIO(data), data_only=True, read_only=True)
    worksheet = workbook["Overzicht"] if "Overzicht" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    return worksheet.title, [list(row) for row in worksheet.iter_rows(values_only=True)]


def normalize_header(value):
    return re.sub(r"[^a-z0-9]+", " ", clean_text(value).lower()).strip()


def parse_grouped_overzicht(sheet_name: str, rows: list[list[object]]):
    group_start_col = {}
    col = 1
    for label in ["Breewel", "ML Express", "De Wit", "De Wit 2", "Totaal"]:
        group_start_col[label] = col
        col += 3

    parsed = []
    for row_index, row in enumerate(rows[2:], start=3):
        label = clean_text(row[0] if len(row) > 0 else "")
        if not label or label.lower().startswith("totaal"):
            continue
        match = DATE_RE.search(label)
        if not match:
            continue
        iso_date = f"{match.group(3)}-{match.group(2)}-{match.group(1)}"
        for group in CARRIER_GROUPS:
            start_col = group_start_col[group["sheetLabel"]]
            dc = to_number(row[start_col] if len(row) > start_col else 0)
            dcs = to_number(row[start_col + 1] if len(row) > start_col + 1 else 0)
            dco = to_number(row[start_col + 2] if len(row) > start_col + 2 else 0)
            if dc == 0 and dcs == 0 and dco == 0:
                continue
            parsed.append({
                "type": "OUT",
                "action_date": iso_date,
                "source_date_label": label,
                "source_row_number": row_index,
                "country": "FR",
                "customer_name": group["siteKlant"],
                "metrics": {
                    "dc": dc,
                    "cctag": 0,
                    "dcs": dcs,
                    "dco": dco,
                    "pal": 0,
                    "vk": 0,
                },
            })
    return {
        "sheet_name": sheet_name,
        "records": parsed,
    }


def parse_export2_rows(sheet_name: str, rows: list[list[object]]):
    if not rows:
        return {"sheet_name": sheet_name, "records": []}
    headers = [normalize_header(value) for value in rows[0]]
    header_index = {name: idx for idx, name in enumerate(headers) if name}

    country_idx = header_index.get("country", 0)
    date_idx = header_index.get("date", 3)
    carrier1_idx = header_index.get("carrier1", 10)
    carrier2_idx = header_index.get("carrier2", 11)
    dc_idx = header_index.get("fustdc", 18)
    dcs_idx = header_index.get("fustdcs", 19)
    dco_idx = header_index.get("fustdco", 20)

    current_country = ""
    current_date = ""
    grouped = {}

    for row_index, row in enumerate(rows[1:], start=2):
        country = clean_text(row[country_idx] if len(row) > country_idx else "")
        if country:
            current_country = country

        date_value = row[date_idx] if len(row) > date_idx else None
        if date_value:
            if hasattr(date_value, "strftime"):
                current_date = date_value.strftime("%Y-%m-%d")
            else:
                date_text = clean_text(date_value)
                match = DATE_RE.search(date_text)
                if match:
                    current_date = f"{match.group(3)}-{match.group(2)}-{match.group(1)}"

        carrier1_raw = clean_text(row[carrier1_idx] if len(row) > carrier1_idx else "")
        carrier2_raw = clean_text(row[carrier2_idx] if len(row) > carrier2_idx else "")
        if not carrier1_raw or not current_date:
            continue
        carrier1_name = CARRIER_NAME_MAP.get(carrier1_raw.lower(), carrier1_raw)
        carrier2_name = CARRIER_NAME_MAP.get(carrier2_raw.lower(), carrier2_raw) if carrier2_raw else ""
        dc = to_number(row[dc_idx] if len(row) > dc_idx else 0)
        dcs = to_number(row[dcs_idx] if len(row) > dcs_idx else 0)
        dco = to_number(row[dco_idx] if len(row) > dco_idx else 0)
        if dc == 0 and dcs == 0 and dco == 0:
            continue

        key = (current_date, current_country or "FR", carrier1_name, carrier2_name)
        if key not in grouped:
            grouped[key] = {
                "type": "OUT",
                "action_date": current_date,
                "source_date_label": current_date,
                "source_row_number": row_index,
                "country": current_country or "FR",
                "customer_name": carrier1_name,
                "carrier1_name": carrier1_name,
                "carrier2_name": carrier2_name,
                "metrics": {
                    "dc": 0,
                    "cctag": 0,
                    "dcs": 0,
                    "dco": 0,
                    "pal": 0,
                    "vk": 0,
                },
            }
        grouped[key]["metrics"]["dc"] += dc
        grouped[key]["metrics"]["dcs"] += dcs
        grouped[key]["metrics"]["dco"] += dco

    return {
        "sheet_name": sheet_name,
        "records": list(grouped.values()),
    }


def parse_rows(input_path: Path):
    sheet_name, rows = load_rows(input_path)
    header_values = [normalize_header(value) for value in (rows[0] if rows else [])]
    if "invoiceweek" in header_values and "carrier1" in header_values and "fustdc" in header_values:
        return parse_export2_rows(sheet_name, rows)
    return parse_grouped_overzicht(sheet_name, rows)


def main():
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command", required=True)
    parse_parser = subparsers.add_parser("parse")
    parse_parser.add_argument("--input", required=True)
    args = parser.parse_args()

    if args.command == "parse":
        payload = parse_rows(Path(args.input))
        print(json.dumps(payload))


if __name__ == "__main__":
    main()
