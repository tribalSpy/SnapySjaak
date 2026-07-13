#!/usr/bin/env python3
import csv
import json
import re
import sys
from io import StringIO
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    from pypdf import PdfReader
except ImportError:
    try:
        from PyPDF2 import PdfReader
    except ImportError:
        PdfReader = None


def clean_text(value):
    return str(value or "").replace("\x00", "").strip()


def split_clean_lines(text):
    return [clean_text(line) for line in str(text or "").splitlines() if clean_text(line)]


def limit_lines(lines, max_lines=220):
    if len(lines) <= max_lines:
      return lines
    kept = lines[:max_lines]
    kept.append(f"... truncated, {len(lines) - max_lines} more lines")
    return kept


def normalize_key(value):
    return re.sub(r"[^a-z0-9]+", " ", clean_text(value).lower()).strip()


def parse_number(value):
    text = clean_text(value).replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_int_like(value):
    number = parse_number(value)
    if number is None:
        return None
    return int(round(number))


def row_find_index(values, expected):
    target = normalize_key(expected)
    for index, value in enumerate(values):
        if normalize_key(value) == target:
            return index
    return None


def row_find_first_index(values, candidates):
    for candidate in candidates:
        found = row_find_index(values, candidate)
        if found is not None:
            return found
    return None


def row_find_label_value(values, expected):
    target = normalize_key(expected)
    for index, value in enumerate(values):
        if normalize_key(value) != target:
            continue
        for next_index in range(index + 1, len(values)):
            candidate = clean_text(values[next_index])
            if candidate:
                return candidate
    return ""


def normalize_known_csi_group(value):
    text = normalize_key(value)
    if not text:
        return ""
    if "cites ge non flowering" in text:
        return "CITES ge. non-flowering p"
    if "other non flowering plant" in text:
        return "Other non-flowering plant"
    if "cites flowering plants" in text or "flowering plants no cactu" in text:
        return "Flowering plants(no cactu"
    if text == "perennials" or "perennials" in text:
        return "Perennials"
    if text == "others" or text.endswith(" others"):
        return "Others"
    if "refined roses" in text:
        return "refined roses"
    if "flowers other fresh" in text:
        return "Flowers (other fresh)"
    if "flowers carnation" in text or "flowers carnations" in text:
        return "Flowers carnation"
    if "flowers chrysanthem" in text:
        return "Flowers chrysanthemums"
    if "flowers roses" in text:
        return "Flowers roses"
    if "flowers lilies" in text:
        return "Flowers lilies"
    if "flowers orchids" in text:
        return "Flowers orchids"
    if "flowers green" in text:
        return "Flowers green"
    return ""


def match_any_token(text, tokens):
    return any(token in text for token in tokens)


def map_ambiguous_plant_group(genus_key):
    if match_any_token(genus_key, ["aloe", "curio", "crassula", "echeveria", "succulent", "cactus", "rhipsalis", "sageretia", "bonsai"]):
        return "CITES ge. non-flowering p"
    if match_any_token(genus_key, ["chlorophytum", "dracaena", "dypsis", "epipremnum", "fittonia", "maranta", "nephrolepis", "schefflera", "spathiphyllum", "zamioculcas", "sansevieria", "sanseveria"]):
        return "Other non-flowering plant"
    return ""


def map_ipaffs_product(genus, commodity_code):
    genus_key = normalize_key(genus)
    code = re.sub(r"\D+", "", clean_text(commodity_code))
    if code.startswith("060240") or code.startswith("60240"):
        return "refined roses"
    if code.startswith("060290500") or code.startswith("60290500") or code.startswith("6029050"):
        return "Perennials"
    if code.startswith("060319700") or code.startswith("60319700") or code.startswith("6031970"):
        return "Others"
    if code.startswith("06029091") or code.startswith("6029091"):
        return "Flowering plants(no cactu"
    known_group = normalize_known_csi_group(genus)
    if known_group in {"refined roses", "Perennials", "Others", "Flowering plants(no cactu"}:
        return known_group
    if code.startswith("060290990") or code.startswith("60290990") or code.startswith("060290991") or code.startswith("60290991"):
        ambiguous_group = map_ambiguous_plant_group(genus_key)
        if ambiguous_group:
            return ambiguous_group
        if known_group in {"CITES ge. non-flowering p", "Other non-flowering plant"}:
            return known_group
    if code.startswith("603140") or code.startswith("060314"):
        return "Flowers chrysanthemums"
    if code.startswith("603120") or code.startswith("060312"):
        return "Flowers carnation"
    if code.startswith("603110") or code.startswith("060311"):
        return "Flowers roses"
    if code.startswith("603150") or code.startswith("060315"):
        return "Flowers lilies"
    if code.startswith("604209"):
        return "Flowers green"
    if "cupressus" in genus_key:
        return "Perennials"
    if "ficus" in genus_key:
        return "Others"
    if "hibiscus" in genus_key:
        return "Flowering plants(no cactu"
    if "chrysanthem" in genus_key and not (code.startswith("06029091") or code.startswith("6029091")):
        return "Flowers chrysanthemums"
    if "dianthus" in genus_key and not (code.startswith("06029091") or code.startswith("6029091")):
        return "Flowers carnation"
    if "gypsoph" in genus_key:
        return "Flowers (other fresh)"
    if "rosa" in genus_key and not (code.startswith("060240") or code.startswith("60240")):
        return "Flowers roses"
    if "solidago" in genus_key:
        return "Flowers (other fresh)"
    if "bonsai" in genus_key or "sageretia" in genus_key or "aloe" in genus_key or "rhipsalis" in genus_key:
        return "CITES ge. non-flowering p"
    if any(token in genus_key for token in ["salvia", "lavandula", "helleborus", "campanula"]):
        return "Perennials"
    if any(token in genus_key for token in ["dypsis", "maranta", "calathea", "chlorophytum", "dracaena", "epipremnum", "fittonia", "nephrolepis", "schefflera", "spathiphyllum", "sansevieria", "sanseveria", "zamioculcas"]):
        return "Other non-flowering plant"
    if any(token in genus_key for token in ["curio", "crassula", "echeveria", "succulent", "cactus"]):
        return "CITES ge. non-flowering p"
    if any(token in genus_key for token in ["echeveria", "fuchsia", "gerbera", "guzmania", "kalanchoe", "phalaenopsis", "anthurium", "celosia", "cymbidium", "cyclamen", "crassula", "begonia", "helianthus", "hydrangea", "mandevilla", "lithodora", "platycodon", "hibiscus"]):
        return "Flowering plants(no cactu"
    if "rosa" in genus_key:
        return "refined roses"
    return "Flowers (other fresh)"


def find_ipaffs_quantity_columns(row):
    unit_index = None
    for index, value in enumerate(row):
        text = normalize_key(value)
        if text in {"pcs", "pc", "pieces", "piece"}:
            unit_index = index
            break
    if unit_index is None:
        return None, None, None
    quantity_index = unit_index - 1 if unit_index - 1 >= 0 else None
    packages_index = None
    for index in range(quantity_index - 1 if quantity_index is not None else unit_index - 1, -1, -1):
        text = normalize_key(row[index])
        if text in {"pk", "box", "boxes", "pakket", "packages"}:
            packages_index = index - 1 if index - 1 >= 0 else None
            break
    return packages_index, quantity_index, unit_index


def parse_ipaffs_rows(rows):
    parsed_rows = []
    summary = {}
    header_values = rows[0] if rows else []
    normalized_header = [normalize_key(cell) for cell in header_values]
    has_header = bool(normalized_header) and any(normalized_header)
    commodity_index = None
    genus_index = None
    packages_index = None
    quantity_index = None
    unit_index = None
    weight_index = None
    if has_header:
        commodity_index = row_find_first_index(header_values, [
            "fullClassificationCode",
            "taricCode",
            "Commodity code",
            "Commodity",
            "classificationType TARIC",
        ])
        if commodity_index is None:
            commodity_index = row_find_index(header_values, "classificationType")
        genus_index = row_find_first_index(header_values, [
            "goodsDescriptionText",
            "classificationValue",
            "Genus",
            "Goods description",
        ])
        if genus_index is None:
            genus_index = row_find_index(header_values, "genus")
        if genus_index is None:
            genus_index = row_find_index(header_values, "product")
        packages_index = row_find_first_index(header_values, ["packages", "Class"])
        quantity_index = row_find_first_index(header_values, ["quantityValue", "for", "Quantity"])
        unit_index = row_find_first_index(header_values, ["quantityUnit", "final", "Unit"])
        weight_index = (
            row_find_index(header_values, "netMassValue")
            if row_find_index(header_values, "netMassValue") is not None
            else row_find_index(header_values, "grossMassValue")
        )
    use_header_indexes = quantity_index is not None or packages_index is not None or commodity_index is not None or genus_index is not None
    start_index = 1 if use_header_indexes or (has_header and any("commodity" in cell or "genus" in cell for cell in normalized_header)) else 0
    for row in rows[start_index:]:
        if not any(clean_text(cell) for cell in row):
            continue
        inferred_packages_index, inferred_quantity_index, inferred_unit_index = find_ipaffs_quantity_columns(row)
        commodity_code = clean_text(row[commodity_index] if commodity_index is not None and commodity_index < len(row) else (row[0] if len(row) > 0 else ""))
        genus_fallback_index = 2 if len(row) > 2 and normalize_key(row[1]) in {"pl", "pl."} else 1
        genus = clean_text(row[genus_index] if genus_index is not None and genus_index < len(row) else (row[genus_fallback_index] if len(row) > genus_fallback_index else ""))
        packages = parse_int_like(
            row[packages_index] if packages_index is not None and packages_index < len(row)
            else (row[inferred_packages_index] if inferred_packages_index is not None and inferred_packages_index < len(row) else (row[7] if len(row) > 7 else ""))
        )
        quantity = parse_int_like(
            row[quantity_index] if quantity_index is not None and quantity_index < len(row)
            else (row[inferred_quantity_index] if inferred_quantity_index is not None and inferred_quantity_index < len(row) else (row[9] if len(row) > 9 else ""))
        )
        unit = clean_text(
            row[unit_index] if unit_index is not None and unit_index < len(row)
            else (row[inferred_unit_index] if inferred_unit_index is not None and inferred_unit_index < len(row) else (row[10] if len(row) > 10 else ""))
        )
        weight = parse_number(row[weight_index] if weight_index is not None and weight_index < len(row) else (row[11] if len(row) > 11 else ""))
        if not commodity_code and not genus and quantity is None and packages is None:
            continue
        product = map_ipaffs_product(genus, commodity_code)
        parsed_row = {
            "product": product,
            "commodity_code": commodity_code,
            "genus": genus,
            "mapped_group": product,
            "packages": packages,
            "quantity": quantity,
            "unit": unit,
            "weight": weight,
        }
        parsed_rows.append(parsed_row)
        summary[product] = int(summary.get(product, 0) + (quantity or 0))
    return {
        "rows": parsed_rows,
        "product_totals": [{"product": product, "quantity": quantity} for product, quantity in summary.items()],
    }


def parse_delimited_rows(text):
    sample = text[:4096]
    delimiters = ",;\t|"
    dialect = None
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=delimiters)
    except Exception:
        dialect = None

    if dialect is not None:
        reader = csv.reader(StringIO(text), dialect)
        rows = list(reader)
        if rows:
            return rows, dialect.delimiter

    fallback_delimiter = ","
    for candidate in [";", "\t", "|", ","]:
        lines = [line for line in text.splitlines() if line.strip()]
        if not lines:
            continue
        split_rows = [line.split(candidate) for line in lines]
        max_columns = max((len(row) for row in split_rows), default=0)
        if max_columns > 1:
            return split_rows, candidate

    reader = csv.reader(StringIO(text))
    return list(reader), fallback_delimiter


def parse_export_sheet(workbook):
    worksheet = workbook.worksheets[0]
    rows = []
    summary = {}
    column_indexes = None
    for row in worksheet.iter_rows(values_only=True):
        values = [clean_text(cell) for cell in row]
        if column_indexes is None:
            description_index = row_find_index(values, "Goods description")
            commodity_index = row_find_index(values, "Commodity code")
            quantity_index = row_find_index(values, "Quantity")
            origin_index = row_find_index(values, "oorsprong")
            if description_index is not None and commodity_index is not None and quantity_index is not None:
                column_indexes = {
                    "description": description_index,
                    "commodity_code": commodity_index,
                    "quantity": quantity_index,
                    "origin": origin_index,
                }
            continue
        if column_indexes is None:
            continue
        description = values[column_indexes["description"]] if column_indexes["description"] < len(values) else ""
        commodity_code = values[column_indexes["commodity_code"]] if column_indexes["commodity_code"] < len(values) else ""
        quantity = parse_int_like(values[column_indexes["quantity"]] if column_indexes["quantity"] < len(values) else "")
        origin_index = column_indexes.get("origin")
        origin = values[origin_index] if origin_index is not None and origin_index < len(values) else ""
        if not description or not commodity_code or quantity is None:
            continue
        parsed_row = {
            "product": description,
            "origin": origin,
            "commodity_code": commodity_code,
            "mapped_group": map_ipaffs_product(description, commodity_code),
            "quantity": quantity,
        }
        rows.append(parsed_row)
        key = f"{description} | {origin}".strip()
        summary[key] = int(summary.get(key, 0) + quantity)
    return {
        "rows": rows,
        "product_origin_totals": [{"product_origin": key, "quantity": quantity} for key, quantity in summary.items()],
    }


def parse_invoice_sheet(workbook):
    worksheet = workbook.worksheets[0]
    rows = []
    summary = {}
    meta = {
        "invoice_number": "",
        "date": "",
        "truck": "",
        "delivery_terms": "",
        "currency": "",
    }
    column_indexes = None
    for row in worksheet.iter_rows(values_only=True):
        values = [clean_text(cell) for cell in row]
        meta["date"] = meta["date"] or row_find_label_value(values, "Date :")
        meta["invoice_number"] = meta["invoice_number"] or row_find_label_value(values, "Invoice nr :")
        meta["truck"] = meta["truck"] or row_find_label_value(values, "Licence Truck :")
        meta["delivery_terms"] = meta["delivery_terms"] or row_find_label_value(values, "Delivery Terms :")
        meta["currency"] = meta["currency"] or row_find_label_value(values, "Currency of invoice")

        if column_indexes is None:
            commodity_index = row_find_index(values, "classificationType TARIC")
            description_index = row_find_index(values, "Goods description")
            origin_index = row_find_index(values, "Origin")
            quantity_index = row_find_index(values, "Quantity")
            if commodity_index is not None and description_index is not None and origin_index is not None and quantity_index is not None:
                column_indexes = {
                    "commodity_code": commodity_index,
                    "description": description_index,
                    "origin": origin_index,
                    "quantity": quantity_index,
                }
            continue
        if column_indexes is None:
            continue
        commodity_code = values[column_indexes["commodity_code"]] if column_indexes["commodity_code"] < len(values) else ""
        description = values[column_indexes["description"]] if column_indexes["description"] < len(values) else ""
        origin = values[column_indexes["origin"]] if column_indexes["origin"] < len(values) else ""
        quantity = parse_int_like(values[column_indexes["quantity"]] if column_indexes["quantity"] < len(values) else "")
        if not description or not commodity_code or quantity is None:
            continue
        parsed_row = {
            "product": description,
            "origin": origin,
            "commodity_code": commodity_code,
            "mapped_group": map_ipaffs_product(description, commodity_code),
            "quantity": quantity,
        }
        rows.append(parsed_row)
        key = f"{description} | {origin}".strip()
        summary[key] = int(summary.get(key, 0) + quantity)
    return {
        "meta": meta,
        "rows": rows,
        "product_origin_totals": [{"product_origin": key, "quantity": quantity} for key, quantity in summary.items()],
    }


def find_line_value(lines, label, max_lookahead=4):
    label_key = normalize_key(label)
    for index, line in enumerate(lines):
        current_key = normalize_key(line)
        if label_key and label_key in current_key:
            tail = line.split(label, 1)[-1].strip(" :") if label in line else ""
            if tail:
                return tail
            for next_index in range(index + 1, min(len(lines), index + 1 + max_lookahead)):
                candidate = clean_text(lines[next_index])
                if candidate and not candidate.startswith("[Page]"):
                    return candidate
    return ""


def parse_temp_phyto_product_lines_from_flat_text(text):
    flattened = re.sub(r"\s+", " ", clean_text(text))
    if not flattened:
        return []

    rows = []
    pattern = re.compile(
        r"(?P<code>\d{3,4})\s+"
        r"(?P<product>.+?)\s+"
        r"(?P<packages>\d+)\s+Box(?:es)?\s+"
        r"(?P<quantity>[\d.,]+)\s+Pieces"
        r"(?=(?:\s+\d{3,4}\s+)|\s*(?:-+\s*<\s*TEXT\s+END|TEXT\s+END|$))",
        flags=re.IGNORECASE,
    )
    for match in pattern.finditer(flattened):
        product = clean_text(match.group("product"))
        if not product:
            continue
        rows.append({
            "product": product,
            "packages": parse_int_like(match.group("packages")),
            "quantity": parse_int_like(match.group("quantity")),
        })
    return rows


def score_temp_phyto_parse(parsed):
    if not isinstance(parsed, dict):
        return -999
    score = 0
    product_lines = parsed.get("product_lines") or []
    if parsed.get("pcnu_number"):
        score += 25
    score += len(product_lines) * 8
    score += sum(2 for line in product_lines if line.get("quantity") is not None)
    if parsed.get("total_quantity") is not None:
        score += 4
    if parsed.get("destination_country"):
        score += 2
    if parsed.get("origin_country"):
        score += 2
    score -= len(parsed.get("problems") or [])
    return score


def parse_temp_phyto_pdf_text(text):
    lines = split_clean_lines(text)
    if not lines:
        return {}

    parsed = {
        "document_state": "ok",
        "pcnu_number": "",
        "destination_country": "",
        "origin_country": "",
        "consignee": "",
        "product_lines": [],
        "total_quantity": None,
        "problems": [],
    }

    lowered = normalize_key(text)
    if "nog niet geactiveerd" in lowered or "not activated" in lowered:
        parsed["document_state"] = "not_activated"
        parsed["problems"].append("Temporary phyto document appears not activated")

    pcnu_match = re.search(r"PCNU\s+([A-Z0-9][A-Z0-9\s-]{5,})", text, flags=re.IGNORECASE)
    if pcnu_match:
        parsed["pcnu_number"] = re.sub(r"[^A-Z0-9]+", "", clean_text(pcnu_match.group(1)).upper())
    else:
        parsed["problems"].append("PCNU number not found")

    parsed["destination_country"] = find_line_value(lines, "to Plant Protection Organization(s) of")
    if not parsed["destination_country"]:
        parsed["problems"].append("Destination country not found")

    parsed["origin_country"] = find_line_value(lines, "Place of origin")
    if not parsed["origin_country"]:
        parsed["problems"].append("Place of origin not found")

    consignee_lines = []
    consignee_started = False
    for line in lines:
        if normalize_key("Declared name and address of consignee") in normalize_key(line):
            consignee_started = True
            continue
        if consignee_started:
            if re.match(r"^\d+\b", line):
                break
            consignee_lines.append(line)
    parsed["consignee"] = ", ".join(consignee_lines[:5]).strip(", ")

    total_match = re.search(r"TOTAL\s+([\d.,]+)\s+Pieces", text, flags=re.IGNORECASE)
    if total_match:
        parsed["total_quantity"] = parse_int_like(total_match.group(1))

    index = 0
    product_chunks = []
    while index < len(lines):
        line = clean_text(lines[index])
        if not re.match(r"^\d{3,4}\b", line):
            index += 1
            continue
        chunk = [line]
        lookahead = index + 1
        while lookahead < len(lines) and len(chunk) < 5:
            candidate = clean_text(lines[lookahead])
            if not candidate:
                lookahead += 1
                continue
            if re.match(r"^\d{3,4}\b", candidate):
                break
            if "text end" in normalize_key(candidate):
                break
            chunk.append(candidate)
            lookahead += 1
        product_chunks.append(" ".join(chunk))
        index = lookahead

    for chunk_text in product_chunks:
        quantity_match = re.search(r"([\d.,]+)\s+Pieces\b", chunk_text, flags=re.IGNORECASE)
        packages_match = re.search(r"\b(\d+)\s+Box\b", chunk_text, flags=re.IGNORECASE)
        quantity = parse_int_like(quantity_match.group(1)) if quantity_match else None
        packages = parse_int_like(packages_match.group(1)) if packages_match else None
        product_text = re.sub(r"^\d{3,4}\s*", "", chunk_text)
        product_text = re.sub(r"\s+\d+\s+Box\b.*$", "", product_text, flags=re.IGNORECASE)
        product_text = re.sub(r"\s+[\d.,]+\s+Pieces\b.*$", "", product_text, flags=re.IGNORECASE)
        cleaned_product = clean_text(product_text)
        if not cleaned_product:
            continue
        parsed["product_lines"].append({
            "product": cleaned_product,
            "packages": packages,
            "quantity": quantity,
        })

    if len(parsed["product_lines"]) <= 1:
        flat_rows = parse_temp_phyto_product_lines_from_flat_text(text)
        if len(flat_rows) > len(parsed["product_lines"]):
            parsed["product_lines"] = flat_rows

    if len(parsed["product_lines"]) == 1 and parsed["product_lines"][0].get("quantity") is None and parsed["total_quantity"] is not None:
        parsed["product_lines"][0]["quantity"] = parsed["total_quantity"]

    if not parsed["product_lines"] and parsed["document_state"] == "ok":
        parsed["problems"].append("No product lines extracted from temporary phyto PDF")

    return parsed


def best_temp_phyto_parse(text_candidates):
    best = {}
    best_score = -999
    for candidate in text_candidates:
        parsed = parse_temp_phyto_pdf_text(candidate)
        score = score_temp_phyto_parse(parsed)
        if score > best_score:
            best = parsed
            best_score = score
    return best


def extract_csv(path: Path):
    try:
        text = path.read_text(encoding="utf-8-sig", errors="replace")
    except Exception:
        text = path.read_text(encoding="utf-8", errors="replace")
    raw_rows, detected_delimiter = parse_delimited_rows(text)
    rows = []
    for row in raw_rows:
        cleaned = [clean_text(cell) for cell in row]
        if any(cleaned):
            rows.append("\t".join(cleaned))
    payload = {
        "content_type": "csv",
        "text": "\n".join(limit_lines(rows)),
        "line_count": len(rows),
        "delimiter": detected_delimiter,
    }
    if raw_rows:
        payload["parsed_data"] = {
            **parse_ipaffs_rows(raw_rows),
            "delimiter": detected_delimiter,
        }
    return payload


def extract_xlsx(path: Path, kind=""):
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to read .xlsx files")
    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    lines = []
    sheet_names = []
    for worksheet in workbook.worksheets:
        sheet_names.append(worksheet.title)
        lines.append(f"[Sheet] {worksheet.title}")
        for row in worksheet.iter_rows(values_only=True):
            cleaned = [clean_text(cell) for cell in row]
            if any(cleaned):
                lines.append("\t".join(cleaned))
    payload = {
        "content_type": "xlsx",
        "sheet_names": sheet_names,
        "text": "\n".join(limit_lines(lines)),
        "line_count": len(lines),
    }
    lower_name = path.name.lower()
    normalized_kind = clean_text(kind)
    if normalized_kind == "generated_invoice" or ("invoice " in lower_name):
        payload["parsed_data"] = parse_invoice_sheet(workbook)
    else:
        payload["parsed_data"] = parse_export_sheet(workbook)
    return payload


def extract_xls(path: Path):
    if xlrd is None:
        raise RuntimeError("xlrd is required to read .xls files")
    workbook = xlrd.open_workbook(path.as_posix())
    lines = []
    sheet_names = []
    for index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(index)
        sheet_names.append(sheet.name)
        lines.append(f"[Sheet] {sheet.name}")
        for row_index in range(sheet.nrows):
            row = [clean_text(sheet.cell_value(row_index, col_index)) for col_index in range(sheet.ncols)]
            if any(row):
                lines.append("\t".join(row))
    return {
        "content_type": "xls",
        "sheet_names": sheet_names,
        "text": "\n".join(limit_lines(lines)),
        "line_count": len(lines),
    }


def extract_pdf(path: Path, kind=""):
    if PdfReader is None:
        return {
            "content_type": "pdf",
            "text": "",
            "line_count": 0,
            "note": "PDF text extractor not installed, keep for manual review",
        }
    reader = PdfReader(str(path))
    lines = []
    temp_phyto_text_candidates = []
    for page_index, page in enumerate(reader.pages, start=1):
        page_variants = []
        for extractor in (
            lambda current_page: current_page.extract_text() or "",
            lambda current_page: current_page.extract_text(extraction_mode="layout") or "",
        ):
            try:
                variant_text = clean_text(extractor(page))
            except TypeError:
                variant_text = ""
            except Exception:
                variant_text = ""
            if variant_text and variant_text not in page_variants:
                page_variants.append(variant_text)
        if not page_variants:
            continue
        default_text = page_variants[0]
        lines.append(f"[Page] {page_index}")
        lines.extend(split_clean_lines(default_text))
        if clean_text(kind).startswith("temp_phyto"):
            for variant_text in page_variants:
                temp_phyto_text_candidates.append(f"[Page] {page_index}\n{variant_text}")
    payload = {
        "content_type": "pdf",
        "text": "\n".join(limit_lines(lines)),
        "line_count": len(lines),
    }
    if clean_text(kind).startswith("temp_phyto"):
        combined_candidates = []
        if temp_phyto_text_candidates:
            combined_candidates.append("\n".join(temp_phyto_text_candidates))
            combined_candidates.extend(temp_phyto_text_candidates)
        if lines:
            combined_candidates.append("\n".join(lines))
        payload["parsed_data"] = best_temp_phyto_parse(combined_candidates)
    return payload


def extract_file(entry):
    path = Path(entry.get("path") or "")
    suffix = path.suffix.lower()
    kind = clean_text(entry.get("kind"))
    if not path.exists():
        return {
            "kind": kind,
            "name": clean_text(entry.get("name")) or path.name,
            "mime_type": clean_text(entry.get("mime_type")),
            "content_type": "missing",
            "error": "File not found",
            "text": "",
        }

    base = {
        "kind": kind,
        "name": clean_text(entry.get("name")) or path.name,
        "mime_type": clean_text(entry.get("mime_type")),
    }
    try:
        if suffix == ".csv":
            return {**base, **extract_csv(path)}
        if suffix == ".xlsx":
            return {**base, **extract_xlsx(path, kind)}
        if suffix == ".xls":
            return {**base, **extract_xls(path)}
        if suffix == ".pdf":
            return {**base, **extract_pdf(path, kind)}
        return {
            **base,
            "content_type": suffix.lstrip(".") or "binary",
            "text": "",
            "line_count": 0,
            "note": "Unsupported file type for CSI extraction",
        }
    except Exception as error:
        return {
            **base,
            "content_type": suffix.lstrip(".") or "unknown",
            "text": "",
            "line_count": 0,
            "error": str(error),
        }


def main():
    command = sys.argv[1] if len(sys.argv) > 1 else ""
    payload = json.loads(sys.stdin.read() or "{}")
    if command != "extract":
        raise RuntimeError(f"Unknown command: {command}")
    files = payload.get("files") or []
    documents = [extract_file(entry) for entry in files]
    print(json.dumps({"documents": documents}, ensure_ascii=False))


if __name__ == "__main__":
    main()
