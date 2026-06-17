from __future__ import annotations

import argparse
import json
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas
import xlrd

PT_PER_CM = 72 / 2.54


def select_sheet_name(sheet_names: list[str]) -> str:
    if "ERP_PASTE" in sheet_names:
        return "ERP_PASTE"
    if "Blad1" in sheet_names:
        return "Blad1"
    return sheet_names[0]


def load_rows(input_path: Path) -> list[list[object]]:
    suffix = input_path.suffix.lower()
    data = input_path.read_bytes()

    if suffix == ".xls":
        workbook = xlrd.open_workbook(file_contents=data)
        sheet_name = select_sheet_name(workbook.sheet_names())
        sheet = workbook.sheet_by_name(sheet_name)
        return [sheet.row_values(index) for index in range(sheet.nrows)]

    workbook = load_workbook(filename=BytesIO(data), data_only=True, read_only=True)
    sheet_name = select_sheet_name(workbook.sheetnames)
    worksheet = workbook[sheet_name]
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def location_prefix(location: str | None) -> str:
    if not location:
        return ""
    return location[:2]


def customer_prefix(customer_code: str | None) -> str:
    if not customer_code:
        return ""
    if customer_code[:1].isdigit():
        return customer_code[:3]
    return customer_code[:2]


def strip_leading_g(location: str) -> str:
    if location[:1].lower() == "g":
        return location[1:].lstrip()
    return location


def parse_halindeling(input_path: Path) -> list[dict[str, str]]:
    rows = load_rows(input_path)
    parsed: list[dict[str, str]] = []
    current_location: str | None = None

    for row in rows:
        location_value = row[0] if len(row) > 0 else None
        customer_value = row[1] if len(row) > 1 else None
        is_header = False

        if isinstance(location_value, str):
            location_text = location_value.strip()
            if (
                not location_text
                or location_text.startswith("Hal:")
                or location_text.startswith("---")
                or location_text.startswith("#")
                or location_text == "Locatie"
            ):
                is_header = True
            else:
                current_location = location_text

        if is_header:
            continue

        if isinstance(customer_value, str) and customer_value.strip() and current_location:
            parsed.append({"location": current_location, "customer": customer_value.strip()})

    return parsed


def inspect_file(input_path: Path) -> dict[str, object]:
    rows = parse_halindeling(input_path)
    if not rows:
        raise ValueError("Geen geldige halindeling-data gevonden")

    loc_prefixes = sorted({location_prefix(item["location"]) for item in rows if location_prefix(item["location"])})
    cust_prefixes = sorted({customer_prefix(item["customer"]) for item in rows if customer_prefix(item["customer"])})
    cust_by_loc: dict[str, list[str]] = {}
    grouped: dict[str, set[str]] = {}
    for item in rows:
        loc = location_prefix(item["location"])
        cust = customer_prefix(item["customer"])
        grouped.setdefault(loc, set()).add(cust)
    for loc, prefixes in grouped.items():
        cust_by_loc[loc] = sorted(prefix for prefix in prefixes if prefix)

    return {
        "locPrefixes": loc_prefixes,
        "custPrefixes": cust_prefixes,
        "custByLoc": cust_by_loc,
        "totalRows": len(rows),
    }


def generate_pdf(input_path: Path, output_path: Path, loc_prefixes: list[str], cust_prefixes: list[str]) -> None:
    rows = parse_halindeling(input_path)
    filtered = []
    for item in rows:
        loc = location_prefix(item["location"])
        cust = customer_prefix(item["customer"])
        if loc_prefixes and loc not in loc_prefixes:
            continue
        if cust_prefixes and cust not in cust_prefixes:
            continue
        filtered.append(item)

    unique_rows: list[dict[str, str]] = []
    seen_customers: set[str] = set()
    for item in filtered:
        customer = item["customer"]
        if customer in seen_customers:
            continue
        seen_customers.add(customer)
        unique_rows.append(item)

    if not unique_rows:
        raise ValueError("Geen klanten gevonden voor deze filters")

    page_width = 10 * PT_PER_CM
    page_height = 15 * PT_PER_CM
    margin = 0.4 * PT_PER_CM
    gap = 0.4 * PT_PER_CM
    location_ratio = 4

    output_path.parent.mkdir(parents=True, exist_ok=True)
    pdf = canvas.Canvas(str(output_path), pagesize=(page_width, page_height))

    def fit_font_size(text: str, max_width: float, max_height: float) -> int:
        size = 1
        while size < 600:
            next_size = size + 1
            text_width = stringWidth(text, "Helvetica-Bold", next_size)
            text_height = next_size
            if text_width > max_width * 0.97 or text_height > max_height * 0.97:
                break
            size = next_size
        return size

    for item in unique_rows:
        location = strip_leading_g(item["location"])
        customer = item["customer"]

        pdf.saveState()
        pdf.translate(0, page_height)
        pdf.rotate(-90)

        width = page_height
        height = page_width
        inner_width = width - (2 * margin)
        available_height = height - (2 * margin) - gap
        customer_height = available_height / (location_ratio + 1)
        location_height = (available_height * location_ratio) / (location_ratio + 1)

        customer_size = fit_font_size(customer, inner_width, customer_height)
        customer_width = stringWidth(customer, "Helvetica-Bold", customer_size)
        customer_text_height = customer_size * 0.72
        customer_y = height - margin - customer_height + ((customer_height - customer_text_height) / 2) - (customer_size * 0.1)
        pdf.setFont("Helvetica-Bold", customer_size)
        pdf.drawString((width - customer_width) / 2, customer_y, customer)

        target_location_size = customer_size * location_ratio
        max_location_size = fit_font_size(location, inner_width, location_height)
        location_size = min(target_location_size, max_location_size)
        location_width = stringWidth(location, "Helvetica-Bold", location_size)
        location_text_height = location_size * 0.72
        location_y = margin + ((location_height - location_text_height) / 2) - (location_size * 0.1)
        pdf.setFont("Helvetica-Bold", location_size)
        pdf.drawString((width - location_width) / 2, location_y, location)

        pdf.restoreState()
        pdf.showPage()

    pdf.save()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command", required=True)

    inspect_parser = subparsers.add_parser("inspect")
    inspect_parser.add_argument("--input", required=True)

    generate_parser = subparsers.add_parser("generate")
    generate_parser.add_argument("--input", required=True)
    generate_parser.add_argument("--output", required=True)
    generate_parser.add_argument("--loc-prefixes-json", default="[]")
    generate_parser.add_argument("--cust-prefixes-json", default="[]")

    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)

    if args.command == "inspect":
        print(json.dumps(inspect_file(input_path), ensure_ascii=True))
        return

    loc_prefixes = [str(item) for item in json.loads(args.loc_prefixes_json or "[]")]
    cust_prefixes = [str(item) for item in json.loads(args.cust_prefixes_json or "[]")]
    generate_pdf(input_path, Path(args.output), loc_prefixes, cust_prefixes)


if __name__ == "__main__":
    main()
