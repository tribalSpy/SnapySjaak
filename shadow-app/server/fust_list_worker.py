import argparse
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side


ROW_BY_CODE = {
    "510": 19,
    "519": 22,
    "520": 25,
    "525": 28,
    "533": 31,
    "544": 34,
    "560": 37,
    "566": 40,
    "577": 43,
    "596": 46,
    "597": 49,
    "CC": 52,
    "CCO": 55,
    "CCS": 58,
    "VK": 61,
}
DATE_CELL = "C12"
CUSTOMER_CELL = "J12"
TOTAL_OK_COLUMN = "N"
TOTAL_BROKEN_COLUMN = "P"
CODE_COLUMN = "B"
CUSTOM_ROW_START = 64
CUSTOM_ROW_STEP = 3
EXPORTER_LABEL_CELL = "B103"
EXPORTER_BLOCK_CELL = "B104"
SIGNATURE_LABEL_CELL = "J103"
SIGNATURE_BOX_START_ROW = 104
SIGNATURE_BOX_END_ROW = 108
SIGNATURE_BOX_START_COL = 10
SIGNATURE_BOX_END_COL = 16


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def to_count(value):
    text = clean_text(value).replace(",", ".")
    if not text:
        return 0
    try:
        return max(0, int(float(text)))
    except ValueError:
        raise ValueError(f"Invalid quantity: {value}")


def load_payload(path: Path):
    payload = json.loads(path.read_text(encoding="utf-8"))
    customer_name = clean_text(payload.get("customer_name"))
    action_date = clean_text(payload.get("action_date"))
    rows = payload.get("rows") or []
    if not customer_name:
        raise ValueError("Customer is required")
    if not action_date:
        raise ValueError("Action date is required")
    if not isinstance(rows, list):
        raise ValueError("Rows must be a list")
    return {
        "customer_name": customer_name,
        "action_date": action_date,
        "exporter": payload.get("exporter") or {},
        "rows": rows,
    }


def generate_workbook(template_path: Path, payload_path: Path, output_path: Path):
    payload = load_payload(payload_path)
    workbook = load_workbook(template_path)
    worksheet = workbook[workbook.sheetnames[0]]

    worksheet[DATE_CELL] = payload["action_date"]
    worksheet[CUSTOMER_CELL] = payload["customer_name"]

    for code, row_number in ROW_BY_CODE.items():
        worksheet[f"{TOTAL_OK_COLUMN}{row_number}"] = 0
        worksheet[f"{TOTAL_BROKEN_COLUMN}{row_number}"] = 0

    custom_rows = []
    for row in payload["rows"]:
        code = clean_text(row.get("code")).upper()
        if not code:
            continue
        row_number = ROW_BY_CODE.get(code)
        if row_number is None:
            custom_rows.append({
                "code": code,
                "total_ok": to_count(row.get("total_ok")),
                "total_broken": to_count(row.get("total_broken")),
            })
            continue
        worksheet[f"{TOTAL_OK_COLUMN}{row_number}"] = to_count(row.get("total_ok"))
        worksheet[f"{TOTAL_BROKEN_COLUMN}{row_number}"] = to_count(row.get("total_broken"))

    for index, row in enumerate(custom_rows):
        row_number = CUSTOM_ROW_START + (index * CUSTOM_ROW_STEP)
        worksheet[f"{CODE_COLUMN}{row_number}"] = row["code"]
        worksheet[f"{TOTAL_OK_COLUMN}{row_number}"] = row["total_ok"]
        worksheet[f"{TOTAL_BROKEN_COLUMN}{row_number}"] = row["total_broken"]

    exporter_name = clean_text((payload.get("exporter") or {}).get("name"))
    exporter_block = clean_text((payload.get("exporter") or {}).get("block"))
    if exporter_name or exporter_block:
        worksheet[EXPORTER_LABEL_CELL] = exporter_name or "Exporter"
        worksheet[EXPORTER_LABEL_CELL].font = Font(bold=True)
        worksheet[EXPORTER_BLOCK_CELL] = exporter_block or exporter_name
        worksheet[EXPORTER_BLOCK_CELL].alignment = Alignment(wrap_text=True, vertical="top")

    worksheet[SIGNATURE_LABEL_CELL] = "Delivery signature"
    worksheet[SIGNATURE_LABEL_CELL].font = Font(bold=True)
    thin_side = Side(style="thin", color="000000")
    for row_number in range(SIGNATURE_BOX_START_ROW, SIGNATURE_BOX_END_ROW + 1):
        for col_number in range(SIGNATURE_BOX_START_COL, SIGNATURE_BOX_END_COL + 1):
            cell = worksheet.cell(row=row_number, column=col_number)
            cell.border = Border(
                left=thin_side if col_number == SIGNATURE_BOX_START_COL else Side(style=None),
                right=thin_side if col_number == SIGNATURE_BOX_END_COL else Side(style=None),
                top=thin_side if row_number == SIGNATURE_BOX_START_ROW else Side(style=None),
                bottom=thin_side if row_number == SIGNATURE_BOX_END_ROW else Side(style=None),
            )
            cell.alignment = Alignment(vertical="center")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main():
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command", required=True)
    generate_parser = subparsers.add_parser("generate")
    generate_parser.add_argument("--template", required=True)
    generate_parser.add_argument("--payload", required=True)
    generate_parser.add_argument("--output", required=True)
    args = parser.parse_args()

    if args.command == "generate":
        generate_workbook(Path(args.template), Path(args.payload), Path(args.output))
        print(json.dumps({"ok": True, "output": args.output}))


if __name__ == "__main__":
    main()
