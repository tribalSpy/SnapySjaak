from __future__ import annotations

import argparse
import csv
import json
import re
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook

PAGE_W = 595.2755905511812
PAGE_H = 841.8897637795277
DRAW_W = PAGE_H
HEADER_MATCH_LIMIT = 12
HAL_ALLOWED_CUSTOMER = re.compile(r"^[A-Za-z0-9#]+$")


def select_sheet_name(sheet_names: list[str]) -> str:
    if "ERP_PASTE" in sheet_names:
        return "ERP_PASTE"
    if "Blad1" in sheet_names:
        return "Blad1"
    if "Sheet1" in sheet_names:
        return "Sheet1"
    return sheet_names[0]


def decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="ignore")


def load_rows(input_path: Path) -> list[list[object]]:
    suffix = input_path.suffix.lower()
    data = input_path.read_bytes()

    if suffix == ".json":
        payload = json.loads(decode_text(data))
        if not isinstance(payload, list):
            raise ValueError("JSON input must be a list of rows")
        return [row if isinstance(row, list) else [row] for row in payload]

    if suffix == ".csv":
        reader = csv.reader(StringIO(decode_text(data)))
        return [list(row) for row in reader]

    if suffix == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(file_contents=data)
        sheet = workbook.sheet_by_name(select_sheet_name(workbook.sheet_names()))
        return [sheet.row_values(index) for index in range(sheet.nrows)]

    workbook = load_workbook(filename=BytesIO(data), data_only=True, read_only=True)
    worksheet = workbook[select_sheet_name(workbook.sheetnames)]
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def header_key(value: object) -> str:
    text = clean_text(value).lower()
    text = text.replace("_", " ").replace("-", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def matches_header(value: object, options: list[str]) -> bool:
    key = header_key(value)
    return any(option in key for option in options)


def first_matching_index(headers: list[object], options: list[str]) -> int:
    for index, value in enumerate(headers):
        if matches_header(value, options):
            return index
    return -1


def find_header_row(rows: list[list[object]], groups: list[list[str]], minimum_hits: int = 2) -> int:
    limit = min(len(rows), HEADER_MATCH_LIMIT)
    best_index = -1
    best_hits = -1
    for row_index in range(limit):
        row = rows[row_index]
        hits = sum(1 for group in groups if first_matching_index(row, group) >= 0)
        if hits > best_hits:
            best_hits = hits
            best_index = row_index
    if best_hits < minimum_hits:
        raise ValueError("Could not find the expected header row")
    return best_index


def to_positive_int(value: object, fallback: int = 1) -> int:
    text = clean_text(value)
    if not text:
        return fallback
    try:
        parsed = int(float(text.replace(",", ".")))
    except ValueError:
        return fallback
    return parsed if parsed > 0 else fallback


def natural_sort_key(value: object) -> list[object]:
    text = clean_text(value)
    return [int(part) if part.isdigit() else part.lower() for part in re.findall(r"[A-Za-z]+|\d+", text)]


def parse_halindeling(input_path: Path) -> dict[str, str]:
    rows = load_rows(input_path)
    current_location = ""
    lookup: dict[str, str] = {}

    for row in rows:
        location_value = row[0] if len(row) > 0 else None
        customer_value = row[1] if len(row) > 1 else None

        location_text = clean_text(location_value)
        if location_text:
            if (
                location_text.startswith("Hal:")
                or location_text.startswith("---")
                or location_text.startswith("#")
                or location_text == "Locatie"
            ):
                continue
            current_location = location_text

        customer_text = clean_text(customer_value)
        if not current_location or not customer_text or not HAL_ALLOWED_CUSTOMER.match(customer_text):
            continue
        lookup.setdefault(customer_text, current_location)

    if not lookup:
        raise ValueError("Geen geldige halindeling-data gevonden")
    return lookup


def parse_planning_rows(input_path: Path) -> list[dict[str, object]]:
    rows = load_rows(input_path)
    header_row_index = find_header_row(
        rows,
        [
            ["klantcode", "klant", "customer code", "customer", "code"],
            ["naam", "klantnaam", "customer name"],
            ["split cc", "aantal denen", "stickers", "aantal", "cc"],
        ],
    )
    headers = rows[header_row_index]
    customer_index = first_matching_index(headers, ["klantcode", "klant", "customer code", "customer", "code"])
    name_index = first_matching_index(headers, ["naam", "klantnaam", "customer name"])
    count_index = first_matching_index(headers, ["split cc", "aantal denen", "stickers", "aantal", "cc"])
    carrier1_index = first_matching_index(headers, ["carrier 1", "eerste carrier", "1e carrier", "vervoerder 1", "auto 1"])
    carrier2_index = first_matching_index(headers, ["carrier 2", "tweede carrier", "2e carrier", "vervoerder 2", "auto 2"])

    parsed = []
    for row in rows[header_row_index + 1:]:
        customer = clean_text(row[customer_index] if customer_index >= 0 and customer_index < len(row) else "")
        if not customer:
            continue
        parsed.append({
            "split": None,
            "count": to_positive_int(row[count_index] if count_index >= 0 and count_index < len(row) else 1),
            "customer": customer,
            "name": clean_text(row[name_index] if name_index >= 0 and name_index < len(row) else ""),
            "carrier1": clean_text(row[carrier1_index] if carrier1_index >= 0 and carrier1_index < len(row) else ""),
            "carrier2": clean_text(row[carrier2_index] if carrier2_index >= 0 and carrier2_index < len(row) else ""),
            "source": "planning",
        })
    if not parsed:
        raise ValueError("No usable rows found in the planning file")
    return parsed


def parse_split_rows(input_path: Path) -> list[dict[str, object]]:
    rows = load_rows(input_path)
    header_row_index = find_header_row(
        rows,
        [
            ["split", "truck"],
            ["klantcode", "klant", "customer code", "customer", "tolocationid"],
            ["naam", "klantnaam", "customer name", "tolocationname"],
        ],
    )
    headers = rows[header_row_index]
    split_index = first_matching_index(headers, ["split", "truck"])
    customer_index = first_matching_index(headers, ["klantcode", "klant", "customer code", "customer", "tolocationid"])
    name_index = first_matching_index(headers, ["naam", "klantnaam", "customer name", "tolocationname"])
    count_index = first_matching_index(headers, ["split cc", "aantal denen", "stickers", "aantal"])
    carrier1_index = first_matching_index(headers, ["carrier 1", "eerste carrier", "1e carrier", "vervoerder 1"])
    carrier2_index = first_matching_index(headers, ["carrier 2", "tweede carrier", "2e carrier", "vervoerder 2", "route", "2e vervoerder", "fixedvehicleid"])

    parsed = []
    for row in rows[header_row_index + 1:]:
        customer = clean_text(row[customer_index] if customer_index >= 0 and customer_index < len(row) else "")
        split_value = clean_text(row[split_index] if split_index >= 0 and split_index < len(row) else "")
        if not customer:
            continue
        parsed.append({
            "split": split_value or None,
            "count": to_positive_int(row[count_index] if count_index >= 0 and count_index < len(row) else 1),
            "customer": customer,
            "name": clean_text(row[name_index] if name_index >= 0 and name_index < len(row) else ""),
            "carrier1": clean_text(row[carrier1_index] if carrier1_index >= 0 and carrier1_index < len(row) else "") or "ML Express",
            "carrier2": clean_text(row[carrier2_index] if carrier2_index >= 0 and carrier2_index < len(row) else ""),
            "source": "split",
        })
    if not parsed:
        raise ValueError("No usable rows found in the split file")
    return parsed


def inspect_source(kind: str, input_path: Path) -> dict[str, object]:
    rows = parse_planning_rows(input_path) if kind == "planning" else parse_split_rows(input_path)
    return {
        "row_count": len(rows),
        "preview": rows[:10],
        "kind": kind,
    }


def _string_width(text: str, font_name: str, font_size: int) -> float:
    from reportlab.pdfbase.pdfmetrics import stringWidth

    return stringWidth(text, font_name, font_size)


def fit_font_size(text: str, font_name: str, max_width: float, start_size: int, min_size: int) -> int:
    size = start_size
    while size > min_size and _string_width(text, font_name, size) > max_width:
        size -= 2
    return max(size, min_size)


def draw_qr(pdf, value: str, x: float, y: float, size: float) -> None:
    from reportlab.graphics import renderPDF
    from reportlab.graphics.barcode import qr
    from reportlab.graphics.shapes import Drawing

    widget = qr.QrCodeWidget(value)
    bounds = widget.getBounds()
    width = bounds[2] - bounds[0]
    height = bounds[3] - bounds[1]
    drawing = Drawing(size, size, transform=[size / width, 0, 0, size / height, 0, 0])
    drawing.add(widget)
    renderPDF.draw(drawing, pdf, x, y)


def draw_sticker(pdf, location: str, truck_label: str | None, counter: str, customer: str, name: str, carrier1: str, carrier2: str) -> None:
    pdf.saveState()
    pdf.translate(PAGE_W, 0)
    pdf.rotate(90)

    pdf.setFont("Helvetica-Bold", 28)
    pdf.drawString(40, 545, location)

    if truck_label:
        truck_font_size = fit_font_size(f"Truck {truck_label}", "Helvetica-Bold", DRAW_W - 80, 24, 14)
        pdf.setFont("Helvetica-Bold", truck_font_size)
        pdf.drawString(40, 510, f"Truck {truck_label}")

    customer_font_size = fit_font_size(customer, "Helvetica-Bold", DRAW_W - 100, 110, 54)
    customer_width = _string_width(customer, "Helvetica-Bold", customer_font_size)
    pdf.setFont("Helvetica-Bold", customer_font_size)
    pdf.drawString((DRAW_W - customer_width) / 2, 485, customer)

    pdf.setFont("Helvetica-Bold", 56)
    counter_width = _string_width(counter, "Helvetica-Bold", 56)
    pdf.drawString(DRAW_W - 40 - counter_width, 510, counter)

    upper_name = name.upper()
    name_font_size = fit_font_size(upper_name, "Helvetica-Bold", DRAW_W - 100, 32, 18)
    name_width = _string_width(upper_name, "Helvetica-Bold", name_font_size)
    pdf.setFont("Helvetica-Bold", name_font_size)
    pdf.drawString((DRAW_W - name_width) / 2, 425, upper_name)

    qr_size = 200
    qr_x = DRAW_W - qr_size - 40
    draw_qr(pdf, customer, qr_x, 110, qr_size)

    carrier_center = (40 + qr_x - 25) / 2
    if carrier1:
        carrier1_size = fit_font_size(carrier1, "Helvetica-Bold", qr_x - 80, 80, 22)
        carrier1_width = _string_width(carrier1, "Helvetica-Bold", carrier1_size)
        pdf.setFont("Helvetica-Bold", carrier1_size)
        pdf.drawString(carrier_center - carrier1_width / 2, 220, carrier1)
    if carrier2:
        carrier2_size = fit_font_size(carrier2, "Helvetica-Bold", qr_x - 80, 80, 22)
        carrier2_width = _string_width(carrier2, "Helvetica-Bold", carrier2_size)
        pdf.setFont("Helvetica-Bold", carrier2_size)
        pdf.drawString(carrier_center - carrier2_width / 2, 80, carrier2)

    pdf.restoreState()
    pdf.showPage()


def make_pdf(rows: list[dict[str, object]], truck_label: str | None, loc_lookup: dict[str, str], out_path: Path) -> tuple[int, list[str]]:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    def sort_key(row: dict[str, object]) -> tuple[object, list[object], str]:
        location = loc_lookup.get(str(row["customer"]), "")
        return (str(row.get("carrier1") or ""), natural_sort_key(location), location)

    pdf = canvas.Canvas(str(out_path), pagesize=A4)
    sticker_count = 0
    missing_locations: list[str] = []
    for row in sorted(rows, key=sort_key):
        customer = str(row["customer"])
        location = loc_lookup.get(customer, "")
        if not location:
            missing_locations.append(customer)
        name = str(row.get("name") or "")
        carrier1 = str(row.get("carrier1") or "")
        carrier2 = str(row.get("carrier2") or "")
        count = to_positive_int(row.get("count"), 1)
        for index in range(1, count + 1):
            draw_sticker(pdf, location, truck_label, f"{index}-{count}", customer, name, carrier1, carrier2)
            sticker_count += 1
    pdf.save()
    return sticker_count, sorted(set(missing_locations))


def safe_file_token(value: object) -> str:
    text = clean_text(value)
    if not text:
        return "overig"
    return re.sub(r"[^A-Za-z0-9._-]+", "_", text).strip("_") or "overig"


def generate_files(hal_input: Path, planning_input: Path | None, split_input: Path | None, output_dir: Path) -> dict[str, object]:
    if not planning_input and not split_input:
        raise ValueError("Upload at least a planning file or a split file first")

    planning_rows = parse_planning_rows(planning_input) if planning_input else []
    split_rows = parse_split_rows(split_input) if split_input else []
    split_customers = {str(row["customer"]) for row in split_rows}

    combined_rows: list[dict[str, object]] = []
    combined_rows.extend(split_rows)
    combined_rows.extend(
        row for row in planning_rows
        if str(row["customer"]) not in split_customers
    )
    if not combined_rows:
        raise ValueError("No sticker rows available to generate")

    loc_lookup = parse_halindeling(hal_input)
    output_dir.mkdir(parents=True, exist_ok=True)

    truck_groups: dict[str, list[dict[str, object]]] = {}
    overig_rows: list[dict[str, object]] = []
    for row in combined_rows:
        split_value = clean_text(row.get("split"))
        if split_value:
            truck_groups.setdefault(split_value, []).append(row)
        else:
            overig_rows.append(row)

    files = []
    missing_locations: list[str] = []

    for split_value in sorted(truck_groups, key=natural_sort_key):
        output_path = output_dir / f"Stickers_truck_{safe_file_token(split_value)}.pdf"
        sticker_count, missing = make_pdf(truck_groups[split_value], split_value, loc_lookup, output_path)
        missing_locations.extend(missing)
        files.append({
            "name": output_path.name,
            "path": str(output_path),
            "split": split_value,
            "customer_count": len(truck_groups[split_value]),
            "sticker_count": sticker_count,
        })

    if overig_rows:
        output_path = output_dir / "Stickers_overig.pdf"
        sticker_count, missing = make_pdf(overig_rows, None, loc_lookup, output_path)
        missing_locations.extend(missing)
        files.append({
            "name": output_path.name,
            "path": str(output_path),
            "split": None,
            "customer_count": len(overig_rows),
            "sticker_count": sticker_count,
        })

    return {
        "files": files,
        "hal_customer_count": len(loc_lookup),
        "combined_row_count": len(combined_rows),
        "missing_locations": sorted(set(missing_locations)),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command", required=True)

    inspect_parser = subparsers.add_parser("inspect-source")
    inspect_parser.add_argument("--kind", choices=["planning", "split"], required=True)
    inspect_parser.add_argument("--input", required=True)

    generate_parser = subparsers.add_parser("generate")
    generate_parser.add_argument("--hal-input", required=True)
    generate_parser.add_argument("--output-dir", required=True)
    generate_parser.add_argument("--planning-input")
    generate_parser.add_argument("--split-input")

    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.command == "inspect-source":
        payload = inspect_source(args.kind, Path(args.input))
        print(json.dumps(payload, ensure_ascii=True))
        return

    payload = generate_files(
        Path(args.hal_input),
        Path(args.planning_input) if args.planning_input else None,
        Path(args.split_input) if args.split_input else None,
        Path(args.output_dir),
    )
    print(json.dumps(payload, ensure_ascii=True))


if __name__ == "__main__":
    main()
